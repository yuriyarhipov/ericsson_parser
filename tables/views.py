import json
from os.path import join, exists
from os import makedirs
from shutil import copy
import tempfile
import xlsxwriter
from zipfile import ZipFile

from django.shortcuts import HttpResponseRedirect
from django.http import HttpResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from table import Table, get_excel
from files.models import Files, SuperFile, CNATemplate, AuditTemplate, LogicalRelation
from files.excel import Excel, PowerAuditExcel, AuditExcel, DistanceExcel
from django.conf import settings
from django.views.decorators.gzip import gzip_page
from openpyxl import load_workbook
from files.distance import Distance
from files.rnd import Rnd

from rest_framework.decorators import api_view
from rest_framework.response import Response


def handle_uploaded_file(files):
    path = settings.STATICFILES_DIRS[0]
    result = []
    for f in files:
        filename = '/xml/'.join([path, f.name])
        destination = open(filename, 'wb+')
        for chunk in f.chunks():
            destination.write(chunk)
        destination.close()
        result.append(filename)
    return result


@gzip_page
def table(request, table_name):
    project = request.project
    if table_name == 'rnd':
        active_file = request.COOKIES.get('active_file')
        if Files.objects.filter(filename__iexact=active_file, project=request.project).exists():
            active_file = Files.objects.filter(filename__iexact=active_file, project=request.project).first()
        else:
            active_file = Files.objects.filter(project=request.project, network__in=['WCDMA', 'LTE']).first()

        if active_file:
            if active_file.network == 'WCDMA':
                table_name = 'rnd_wcdma'
            else:
                table_name = 'rnd_lte'
            filename = active_file.filename

    current_table = Table(table_name, project.id)
    data = current_table.get_data()
    columns = current_table.columns

    if request.GET.get('excel'):
        return HttpResponseRedirect(Excel(request.project.project_name, table_name, columns, data).filename)

    return HttpResponse(json.dumps({'columns': columns, 'data': data}), content_type='application/json')


def rnd(request):
    columns = []
    data = []
    active_file = request.COOKIES.get('active_file')
    if Files.objects.filter(filename=active_file, project=request.project).exists():
        active_file = Files.objects.filter(filename=active_file, project=request.project).first()
    else:
        active_file = Files.objects.filter(project=request.project, file_type='xml', network_in=['WCDMA', 'LTE']).first()

    if active_file:
        if active_file.network == 'WCDMA':
            table_name = 'rnd_wcdma'
        else:
            table_name = 'rnd_lte'
        current_table = Table(table_name, active_file.filename)
        columns = current_table.columns
        data = current_table.get_data()
    return HttpResponse(json.dumps({'columns': columns, 'data': data}), content_type='application/json')


def explore(request, filename):
    project = request.project
    tree_file = None
    tables = []

    if Files.objects.filter(project=project, filename=filename).exists():
        tree_file = Files.objects.filter(project=project, filename=filename).first()
        if tree_file.network == 'GSM':
            if tree_file.file_type == 'GSM BSS CNA  OSS FILE':
                tables = [{'table': cna_tables.table_name, 'filename': filename} for cna_tables in CNATemplate.objects.all()]
        else:
            tables = [{'table': table, 'filename': filename} for table in request.wcdma.tables.split(',')]
    elif (filename == 'nokia'):
        f = Files.objects.filter(project=project, vendor='Nokia').first()
        for t in f.tables.split(','):
            tables.append({'table': t, 'filename': f.filename})

    tables.sort()
    return HttpResponse(json.dumps(tables), content_type='application/json')


def by_technology(request, network):
    data = []
    file_filter = request.GET.get('filter')
    project = request.project
    if network == 'GSM':
        filename = request.gsm.filename
        for cna_table in CNATemplate.objects.all().order_by('table_name'):
            data.append({'label': cna_table.table_name, 'table': cna_table.table_name, 'type': 'CNA Table', 'filename': filename})

    elif network == 'WCDMA':
        filename = request.wcdma.filename
        tables = request.wcdma.tables.split(',')
        tables.sort()
        data = [
            {'label': 'Universal3g3gNeighbors', 'table': 'Universal3g3gNeighbors', 'type': 'Universal table', 'filename': filename},
            {'label': '3G3GNeighbors', 'table': 'new3g', 'type': 'Additional table', 'filename': filename},
            {'label': 'Topology', 'table': 'topology', 'type': 'Additional table', 'filename': filename},
            {'label': 'RND', 'table': 'rnd_wcdma', 'type': 'Additional table', 'filename': filename},
            {'label': 'BrightcommsRNDDate', 'table': 'BrightcommsRNDDate', 'type': 'Additional table', 'filename': filename},
            {'label': '3GNeighbors', 'table': 'threegneighbors', 'type': 'Additional table', 'filename': filename},
            {'label': '3GIRAT', 'table': '3girat', 'type': 'Additional table', 'filename': filename},
            {'label': '3GMAP_INTRAFREQ', 'table': 'map_intrafreq', 'type': 'Additional table', 'filename': filename},
            {'label': '3GMAP_INTERFREQ', 'table': 'map_interfreq', 'type': 'Additional table', 'filename': filename},
            {'label': '3GMAP_GSMIRAT', 'table': 'map_gsmirat', 'type': 'Additional table', 'filename': filename},
            {'label': '3GNeighbors_CO-SC', 'table': 'neighbors_co-sc', 'type': 'Additional table', 'filename': filename},
            {'label': '3GNeighbors_TWO_WAYS', 'table': 'neighbors_two_ways', 'type': 'Additional table', 'filename': filename},
        ]
        data.extend([{'label': table, 'table': table, 'type': 'XML table', 'filename': filename} for table in tables])

    elif network == 'LTE':
        filename = request.lte.filename
        tables = request.lte.tables.split(',')
        tables.sort()
        data = [
            {'label': 'Topology', 'table': 'topology_lte', 'type': 'Additional table', 'filename': filename},
            {'label': 'RND', 'table': 'rnd_lte', 'type': 'Additional table', 'filename': filename},
        ]
        data.extend([{'label': table, 'table': table, 'type': 'XML table', 'filename': filename} for table in tables])

    data_result = []

    for f in data:
        if file_filter and f.get('label'):
            if file_filter.lower() in f.get('label').lower():
                data_result.append(f)
        else:
            data_result.append(f)

    return HttpResponse(
        json.dumps(data_result),
        content_type='application/json')


def maps(request):
    project = request.project
    filenames = [f.filename for f in Files.objects.filter(project=project)]
    return HttpResponse(
        json.dumps(filenames),
        content_type='application/json')


def map(request, filename):
    project = request.project
    f = Files.objects.filter(project=project, filename=filename).first()
    data = f.get_map()
    return HttpResponse(
        json.dumps(data),
        content_type='application/json')


def set_audit_template(request):
    project = request.project
    network = request.POST.get('network')
    AuditTemplate.objects.filter(project=project, network=network).delete()
    filename = handle_uploaded_file(request.FILES.getlist('uploaded_file'))[0]
    wb = load_workbook(filename=filename, use_iterators=True)
    result = []
    for sheet_name in wb.get_sheet_names():
        ws = wb.get_sheet_by_name(sheet_name)
        for row in ws.iter_rows():
            if (row[0].value != 'Parameter') and (row[1].value):
                result.append({'param': row[0].value, 'value': row[1].value})
                AuditTemplate.objects.create(project=project, network=network, param=row[0].value, value=row[1].value)

    return HttpResponse(json.dumps(result), content_type='application/json')


def get_audit_template(request):
    project = request.project
    network = request.GET.get('network')
    result = []
    for at in AuditTemplate.objects.filter(
            project=project,
            network__iexact=network):
        result.append({'param': at.param, 'value': at.value})
    return HttpResponse(json.dumps(result), content_type='application/json')


def run_audit(request, network, filename):
    project = request.project
    result = []
    total = None
    for at in AuditTemplate.objects.filter(project=project, network=network):
        value = at.check_param(filename)
        complaint = value.get('complaint', 0)
        if not total:
            total = at.total(filename)
        not_complaint = total - complaint
        if total != 0:
            percent = int(float(complaint) / float(total) * 100)
        else:
            percent = 0

        result.append({
            'param': at.param,
            'recommended': at.value,
            'complaint': complaint,
            'not_complaint': not_complaint,
            'total': total,
            'percent': percent})
    chart = []
    for param in result:
        chart.append([param.get('param'), param.get('complaint')])
    return HttpResponse(json.dumps({'table': result, 'chart': chart}), content_type='application/json')


def audit_param(request, network, filename, param_name):
    project = request.project
    result = []
    chart = []
    for row in AuditTemplate.objects.get(project=project, network=network, param=param_name).audit_param(filename):
        pass
    return HttpResponse(json.dumps({'table': result, 'chart': chart}), content_type='application/json')



def excel_audit(request, network, filename):
    project = request.project
    result = []
    for at in AuditTemplate.objects.filter(project=project, network=network):
        value = at.check_param(filename)
        complaint = value.get('complaint', 0)
        not_complaint = value.get('not_complaint', 0)
        total = complaint + not_complaint
        if total != 0:
            percent = int(float(complaint) / float(total) * 100)
        else:
            percent = 0

        result.append({
            'param': at.param,
            'recommended': at.value,
            'complaint': complaint,
            'not_complaint': not_complaint,
            'total': total,
            'percent': percent})
    chart = []
    for param in result:
        chart.append([param.get('param'), param.get('complaint')])

    ae = AuditExcel()
    f = ae.create_file(result, project.project_name, filename)
    return HttpResponseRedirect(f)


def power_audit(request, filename):
    project = request.project
    audit = Files.objects.get(project=project, filename=filename).power_audit()
    chart = [audit.get('sector_count'), len(audit.get('miss_sectors', []))]

    table = audit.get('miss_sectors')
    return HttpResponse(json.dumps({'chart': chart, 'table': table}), content_type='application/json')


def excel_power_audit(request, filename):
    project = request.project
    audit = Files.objects.get(project=project, filename=filename).power_audit()
    pa = PowerAuditExcel()
    f = pa.create_file(audit, project.project_name, filename)
    return HttpResponseRedirect(f)


def get_sectors(request):
    project = request.project
    sectors = Distance().get_sectors(project.id)
    return HttpResponse(
        json.dumps({'sectors': sectors}),
        content_type='application/json')


def get_logical_sectors(request):
    sectors = Distance().get_sectors()
    result = set()
    for s in sectors:
        result.add(s[-1])
    result = list(result)
    result.sort()
    return HttpResponse(
        json.dumps({'sectors': result}),
        content_type='application/json')


def get_rbs(request):
    project = request.project
    rbs = Distance().get_rbs(project.id)
    return HttpResponse(
        json.dumps({'rbs': rbs}),
        content_type='application/json')


def get_dates(request):
    project = request.project
    dates = Distance().get_dates(project.id)
    return HttpResponse(json.dumps(dates), content_type='application/json')


def get_charts(request, date_from, date_to, rbs):
    project = request.project
    charts = Distance().get_charts(date_from, date_to, rbs, project.id)
    return HttpResponse(json.dumps(charts), content_type='application/json')


def get_low_coverage(request, date_from, date_to, distance):
    project = request.project
    coverage = Distance().get_low_coverage(date_from, date_to, int(distance), project.id)
    return HttpResponse(json.dumps(coverage), content_type='application/json')


def get_over_coverage(request, date_from, date_to, distance):
    project = request.project
    coverage = Distance().get_over_coverage(date_from, date_to, int(distance), project.id)
    return HttpResponse(json.dumps(coverage), content_type='application/json')


def get_chart(request, date_from, date_to, sector):
    chart, table, title, distances = Distance().get_chart(date_from, date_to, sector)
    return HttpResponse(
        json.dumps({
            'chart': chart,
            'table': table,
            'title': title,
            'distances': distances}),
        content_type='application/json')


def get_load_distr(request, day_from, day_to, rbs):
    project = request.project
    return HttpResponse(
        json.dumps(Distance().get_distr(day_from, day_to, rbs, project.id)),
        content_type='application/json')


def get_distance_excel(request, date, sector):
    project = request.project
    filename = '%s_%s' % (date, sector)
    chart, table = Distance().get_chart(date, sector)
    f = DistanceExcel().create_file(project.project_name, filename, chart, table)
    return HttpResponseRedirect(f)


@api_view(['GET', 'POST', 'DELETE', ])
def logical_sectors(request, logical_sector=None, sector=None):
    project = request.project
    if request.method == 'POST':
        logical_sector = request.POST.get('logical_sector')
        band = request.POST.get('band')
        technology = request.POST.get('technology')
        sector = request.POST.get('sector')
        Distance().add_logical_sector(project.id, logical_sector, technology, band, sector)

    elif request.method == 'DELETE':
        Distance().delete_logical_sectors(project.id, logical_sector, sector)

    return Response(Distance().logical_sectors(project.id))


@api_view(['GET', ])
def psc_distance(request):
    project = request.project
    filename = request.wcdma.filename
    return Response(Distance().get_psc_distance(project.id, filename))


@api_view(['POST', ])
def save_logical_relation(request):
    project = request.project

    technology_from = request.POST.get('technology_from')
    carrier_from = request.POST.get('carrier_from')
    symmetry = request.POST.get('symmetry')
    technology_to = request.POST.get('technology_to')
    carrier_to = request.POST.get('carrier_to')
    hor_from = ''
    hor_to = ''

    if technology_from == 'GSM':
        hor_from = 'GSM-SCTYPE:%s' % carrier_from
    else:
        hor_from = '%s-Carrier:%s' % (technology_from, carrier_from)

    if technology_to == 'GSM':
        hor_to = 'GSM-SCTYPE:%s' % carrier_to
    else:
        hor_to = '%s-Carrier:%s' % (technology_to, carrier_to)

    message = ''

    if (symmetry == '<---->') and (not LogicalRelation.objects.filter(project=project, technology_from= technology_from, carrier_from=carrier_from, technology_to=technology_to, carrier_to=carrier_to).exists()) and (not LogicalRelation.objects.filter(project=project, technology_from= technology_to, carrier_from=carrier_to, technology_to=technology_from, carrier_to=carrier_from).exists()):
        LogicalRelation.objects.create(
            project=project,
            technology_from= technology_from,
            carrier_from=carrier_from,
            symmetry=symmetry,
            technology_to=technology_to,
            carrier_to=carrier_to,
            ho_realtion = '%s---->%s' % (hor_from, hor_to))
        if (technology_from != technology_to) or (carrier_from != carrier_to):
            LogicalRelation.objects.create(
                project=project,
                technology_from= technology_to,
                carrier_from=carrier_to,
                symmetry=symmetry,
                technology_to=technology_from,
                carrier_to=carrier_from,
                ho_realtion = '%s---->%s' % (hor_to, hor_from))
        message = 'Add logical relation'
    elif (symmetry == '---->') and (not LogicalRelation.objects.filter(project=project, technology_from= technology_from, carrier_from=carrier_from, technology_to=technology_to, carrier_to=carrier_to).exists()):
        LogicalRelation.objects.create(
            project=project,
            technology_from= technology_from,
            carrier_from=carrier_from,
            symmetry=symmetry,
            technology_to=technology_to,
            carrier_to=carrier_to,
            ho_realtion = '%s---->%s' % (hor_from, hor_to))
        message = 'Add logical relation'
    else:
        message = 'Logical Relation with same technology and carrier exists'

    data = []
    for lr in LogicalRelation.objects.filter(project=project):
        data.append(dict(
            id= lr.id,
            ho_realtion= lr.ho_realtion,
        ))
    return Response({'data': data, 'message': message})


@api_view(['GET', ])
def logical_relations(request):
    project = request.project
    data = []
    for lr in LogicalRelation.objects.filter(project=project):
        data.append(dict(
            id= lr.id,
            ho_realtion= lr.ho_realtion,
        ))
    return Response(data)

@api_view(['POST', ])
def delete_logical_relation(request):
    project = request.project
    id = request.POST.get('id')
    data = []
    LogicalRelation.objects.filter(project=project, id=id).delete()
    for lr in LogicalRelation.objects.filter(project=project):
        data.append(dict(
            id= lr.id,
            ho_realtion= lr.ho_realtion,
        ))
    return Response(data)


@api_view(['GET', ])
def psc_collision(request):
    project = request.project
    filename = request.wcdma.filename
    data = Rnd(project.id, 'WCDMA').psc_collision(filename)
    if request.GET.get('excel'):
        excel_data = []
        for r in data:
            excel_data.append([
                r.get('Source'),
                r.get('Label'),
                r.get('Target'),
                r.get('PSC_Source'),
                r.get('PSC_Target'),
                r.get('uarfcnDl_Source'),
                r.get('uarfcnDl_Target'),
                '1st Order Collision'])

        static_path = settings.STATICFILES_DIRS[0]
        file_path = '%s/%s' % (static_path, project.id)
        if not exists(file_path):
            makedirs(file_path)
        excel_filename = join(tempfile.mkdtemp(), 'psc_collision.xlsx')
        workbook = xlsxwriter.Workbook(excel_filename)
        worksheet = workbook.add_worksheet()

        worksheet.insert_image('A1', static_path + '/XMART.png', {
            'x_scale': 0.17,
            'y_scale': 0.2})
        format = workbook.add_format({'font_size': 20, 'bg_color': '#31859c', 'font_color': 'white'})
        worksheet.write(0, 0, '', format)
        worksheet.write(0, 1, '', format)
        worksheet.write(0, 2, '1stOrder_CoSC_Collision', format)
        worksheet.write(0, 3, '', format)
        worksheet.write(0, 4, '', format)
        worksheet.write(0, 5, '', format)
        worksheet.write(0, 6, '', format)
        worksheet.write(0, 7, '', format)
        format_columns = workbook.add_format({'bg_color': '#31859c', 'font_color': 'white'})
        worksheet.write(1, 0, 'Source', format_columns)
        worksheet.write(1, 1, 'Label', format_columns)
        worksheet.write(1, 2, 'Target', format_columns)
        worksheet.write(1, 3, 'PSC_Source', format_columns)
        worksheet.write(1, 4, 'PSC_Target', format_columns)
        worksheet.write(1, 5, 'uarfcnDl_Source', format_columns)
        worksheet.write(1, 6, 'uarfcnDl_Target', format_columns)
        worksheet.write(1, 7, 'Report', format_columns)
        worksheet.set_row(0, 60)
        worksheet.add_table(1, 0, len(excel_data)+1, 7,
                            {'data': excel_data,
                             'header_row': False})
        worksheet.set_column('A:A', 10)
        worksheet.set_column('A:B', 10)
        worksheet.set_column('A:C', 10)
        worksheet.set_column('A:D', 15)
        worksheet.set_column('A:E', 15)
        worksheet.set_column('A:F', 15)
        worksheet.set_column('A:G', 15)
        worksheet.set_column('A:H', 15)
        workbook.close()
        archive_filename = join(file_path, 'psc_collision.zip')
        zip = ZipFile(archive_filename, 'w')
        zip.write(excel_filename, arcname='psc_collision.xlsx')
        zip.close()

        return HttpResponseRedirect(join('/static/%s' % project.id, 'psc_collision.zip'))

    return Response(data)
