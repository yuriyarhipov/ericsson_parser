import json
from os.path import basename

from openpyxl import load_workbook

from django.http import HttpResponse
from django.conf import settings
from django.shortcuts import HttpResponseRedirect

from rest_framework.decorators import api_view
from rest_framework.response import Response

from celery.result import AsyncResult

from .models import Files, UploadedFiles, CNATemplate, FileTasks, DriveTestLegend, WNCS
from tables.table import Table
from files.compare import CompareFiles, CompareTable
from files.excel import ExcelFile, Excel
from files.distance import Distance
from files.drive_test import DriveTest
from project.models import UserSettings, MapUserPosition
from rnd import Rnd
import tasks
from os import makedirs
from os.path import exists
from shapely.geometry import box, Point
from django.views.decorators.gzip import gzip_page
from tables.universal_tables import UniversalTable
from django.contrib.auth.models import User


def handle_uploaded_file(files):
    path = settings.STATICFILES_DIRS[0]
    result = []
    for f in files:
        if not exists(path + '/xml/'):
            makedirs(path + '/xml/')
        filename = '/xml/'.join([path, f.name])
        destination = open(filename, 'wb+')
        for chunk in f.chunks():
            destination.write(chunk)
        destination.close()
        result.append(filename)
    return result


def status(request, id):
    if FileTasks.objects.filter(task_id=id):
        ft = FileTasks.objects.get(task_id=id)
        tasks = ft.tasks.split(',')
        active_tasks = []
        for task_id in tasks:
            if not AsyncResult(task_id).ready():
                active_tasks.append(task_id)
        current = ft.max_value - len(active_tasks)
        value = float(current) / float(ft.max_value) * 100
        return HttpResponse(json.dumps({'current': int(value), }), mimetype='application/json')

    job = AsyncResult(id)
    data = job.result or job.state

    try:
        json_data = json.dumps(data)
    except:
        data = {'current': 100, }
        json_data = json.dumps(data)

    return HttpResponse(json_data, mimetype='application/json')


def uploaded_files(request):
    project = request.project
    data = []
    for f in UploadedFiles.objects.filter(project=project):
        data.append(dict(
            id=f.id,
            filename=basename(f.filename),
            file_type=f.file_type,
            description=f.description,
            vendor=f.vendor,
            network=f.network,
            status=f.status,
            label=f.label,
            date=f.date.strftime('%d.%m.%Y'),
        ))
    return HttpResponse(json.dumps(data), content_type='application/json')

def delete_uploaded_all(request):
    project = request.project
    UploadedFiles.objects.filter(project=project).delete()
    return HttpResponse(json.dumps([]), content_type='application/json')

def delete_uploaded(request):
    project = request.project
    id = request.POST.get('id')
    UploadedFiles.objects.filter(id=id).delete()
    data = []
    for f in UploadedFiles.objects.filter(project=project):
        data.append(dict(
            id=f.id,
            filename=basename(f.filename),
            file_type=f.file_type,
            description=f.description,
            vendor=f.vendor,
            network=f.network,
            date=f.date.strftime('%d.%m.%Y'),
        ))
    return HttpResponse(json.dumps(data), content_type='application/json')


def save_files(request):
    project = request.project
    description = request.POST.get('description')
    vendor = request.POST.get('vendor')
    file_type = request.POST.get('file_type')
    network = request.POST.get('network')
    filenames = handle_uploaded_file(request.FILES.getlist('file'))
    for f in filenames:
        UploadedFiles.objects.create(
            filename=f,
            file_type=file_type,
            description=description,
            vendor=vendor,
            network=network,
            project=project
        )

    data = []
    for f in UploadedFiles.objects.filter(project=project):
        data.append(dict(
            id=f.id,
            filename=basename(f.filename),
            file_type=f.file_type,
            description=f.description,
            vendor=f.vendor,
            network=f.network,
        ))
    return HttpResponse(json.dumps(data), content_type='application/json')


def run_tasks_all(request):
    project = request.project
    data = []    
    for f in UploadedFiles.objects.filter(project=project):
        if f.file_type not in ['RND', 'Drive Test', 'HISTOGRAM FILE COUNTER - Access Distance']:
            for f in Files.objects.filter(project=project, file_type=f.file_type):
                f.clear_tables()
                f.delete()    
    for f in UploadedFiles.objects.filter(project=project):        
        job = tasks.worker.delay(f.filename, project, f.description, f.vendor, f.file_type, f.network, f.id)
    return HttpResponse(json.dumps([]), content_type='application/json')


def files(request):
    project = request.project
    files = []

    active_files = []

    if request.lte:
        active_files.append(request.lte.filename)
    if request.cna:
        active_files.append(request.cna.filename)
    if request.wcdma:
        active_files.append(request.wcdma.filename)

    for f in Files.objects.filter(project=project):
        status = 'uploaded'
        files.append({
            'id': f.id,
            'filename': f.filename,
            'date': f.date.strftime('%m.%d.%Y'),
            'file_type': f.file_type,
            'network': f.network,
            'vendor': f.vendor,
            'description': f.description,
            'status': status
        })
    return HttpResponse(json.dumps({'files': files, }), content_type='application/json')


def measurements(request, file_type):
    project = request.project
    data = []
    for f in Files.objects.filter(project=project, file_type=file_type):
        data.append({'filename': f.filename, 'file_type': f.file_type, 'network': f.network})
    return HttpResponse(json.dumps(data), content_type='application/json')

def licenses(request):
    project = request.project
    data = []
    for f in Files.objects.filter(project=project, file_type__in=['WCDMA LICENSE FILE OSS XML', 'LTE LICENSE FILE OSS XML']):
        data.append({'filename': f.filename, 'file_type': f.file_type, 'network': f.network})
    return HttpResponse(json.dumps(data), content_type='application/json')

def license(request, filename, table):
    current_table = Table(table, filename)
    columns = current_table.columns
    data = current_table.get_data()
    data = data[:20]
    return HttpResponse(json.dumps({'columns': columns, 'data': data}), content_type='application/json')

def hardwares(request):
    project = request.project
    data = []
    for f in Files.objects.filter(project=project, file_type__in=['WCDMA HARDWARE FILE OSS XML', 'LTE HARDWARE FILE OSS XML']):
        data.append({'filename': f.filename, 'file_type': f.file_type, 'network': f.network})
    return HttpResponse(json.dumps(data), content_type='application/json')

def get_files_for_compare(request, network):
    main_file = ''
    tables = []

    if network == 'WCDMA':
        main_file = request.wcdma.filename
        for table in request.wcdma.tables.split(','):
            tables.append({'table': table, 'type': 'Table'})

    data = [file.filename for file in Files.objects.filter(project=request.project, network=network, file_type='xml')]
    data.sort()
    return HttpResponse(json.dumps({'files': data, 'main_file': main_file, 'tables': tables}), content_type='application/json')

def compare_files(request):
    data = dict()
    filename = request.POST.get('main_file')
    files = request.POST.getlist('files')
    table = request.POST.get('table')
    cells = request.POST.getlist('cells')
    if not table:
        data['compare_files'] = CompareFiles(filename, files).get_tables_info()
    if (not cells) and table:
        ct = CompareTable(table, filename, files)
        mo = ct.get_mo_list()
        data['compare_table'] = {'columns': ct.columns, 'data': ct.get_data(mo)}

    return HttpResponse(json.dumps(data), content_type='application/json')

def delete_file(request, id):
    project = request.project
    f = Files.objects.get(id=id)
    Distance().delete_file(f.filename, project.id)
    f.delete()
    # SuperFile.objects.filter(filename=f.filename).delete()
    return HttpResponse(json.dumps([]), content_type='application/json')


def get_files(request, network):
    project = request.project
    data = [f.filename for f in Files.objects.filter(project=project, network=network)]
    return HttpResponse(json.dumps(data), content_type='application/json')


def save_superfile(request):
    project = request.project
    filename = request.POST.get('filename')
    selected_files = request.POST.getlist('files')
    source_file = Files.objects.filter(filename=selected_files[0], project=project).first()
    network = source_file.network
    file_type = source_file.file_type
    vendor = source_file.vendor
    tasks.superfile.delay(filename, selected_files)
    #SuperFile.objects.filter(filename=filename, network=network, project=project).delete()
    #SuperFile.objects.create(
    #    filename=filename,
    #    files=','.join(selected_files),
    #    network=network,
    #    project=project,
    #    file_type=file_type,
    #    vendor=vendor)
    return HttpResponse(json.dumps({'status': 'ok'}), content_type='application/json')


def set_cna_template(request):
    project = request.project
    CNATemplate.objects.filter(project=project).delete()

    filename = handle_uploaded_file(request.FILES.getlist('uploaded_file'))[0]
    wb = load_workbook(filename=filename)

    for sheet_name in wb.get_sheet_names():
        ws = wb.get_sheet_by_name(sheet_name)
        columns = []
        for col in ws.columns:
            if col[0].value:
                columns.append(col[0].value)
        sql_columns = ','.join(columns)

        table_name = sheet_name.split('-')
        table_name = table_name[len(table_name)-1]
        table_name = table_name.strip().replace(' ', '_')
        CNATemplate.objects.create(project=project, table_name=table_name, columns=sql_columns)

    return HttpResponse(json.dumps([]), content_type='application/json')


def get_cna_template(request):
    tables = []
    project = request.project
    for cna_template in CNATemplate.objects.filter(project=project).order_by('table_name'):
        tables.append({'table_name': cna_template.table_name, 'columns': cna_template.columns})
    return HttpResponse(json.dumps(tables), content_type='application/json')


def get_excel(request, vendor, network, filename):
    tables =  request.GET.getlist('table') 
    ExcelFile(request.project, vendor, network, filename, tables)    
    return HttpResponseRedirect('/static/' + filename + '.zip')


@api_view(['GET', 'POST'])
@gzip_page
def rnd(request, network=None):
    project = request.project
    filenames = request.GET.getlist('filename')
    data = Rnd(project.id, network).get_data(filenames)
    return Response(data)


@api_view(['POST', ])
def map_frame(request, network):
    project = request.project
    data = []
    map_bounds = request.POST.get('bounds').split(',')
    map_box = box(
        float(map_bounds[1]),
        float(map_bounds[0]),
        float(map_bounds[3]),
        float(map_bounds[2]))

    for point in Rnd(project.id, network).get_data().get('data'):
        map_p = Point(point.get('Latitude'), point.get('Longitude'))
        if map_p.within(map_box):
            data.append(point)
    return Response(data)


@api_view(['GET', ])
def init_map(request):
    project = request.project
    username = request.COOKIES.get('username')
    if UserSettings.objects.filter(current_project=project, user=username).exists():
        us = UserSettings.objects.get(current_project=project, user=username)
        lat, lng = us.map_center.split(',')
        data = {'point': [lat, lng], 'zoom': us.map_zoom}
        return Response(data)
        
    filenames = request.GET.getlist('filename')
    gsm_data = Rnd(project.id, 'GSM').get_data(filenames).get('data')
    data = {}
    if gsm_data:
        for s in gsm_data:
            lat = s.get('Latitude')
            lng = s.get('Longitude')
            if (lat and lng):
                data = {'point': [lat, lng]}
                return Response(data)

    wcdma_data = Rnd(project.id, 'WCDMA').get_data(filenames).get('data')
    if wcdma_data:
        for s in wcdma_data:
            lat = s.get('Latitude')
            lng = s.get('Longitude')
            if (lat and lng):
                data = {'point': [lat, lng]}
                return Response(data)

    lte_data = Rnd(project.id, 'LTE').get_data(filenames).get('data')
    if lte_data:
        for s in lte_data:
            lat = s.get('Latitude')
            lng = s.get('Longitude')
            if (lat and lng):
                data = {'point': [lat, lng]}
                return Response(data)
    return Response(data)


@api_view(['POST', ])
def rnd_table(request, network=None):
    project = request.project
    data = []
    if request.method == 'POST':
        Rnd(project.id, network.lower()).save_row(request.POST)
    return Response(data)

@api_view(['GET', ])
def get_user_position(request):
    project = request.project
    username = request.COOKIES.get('username')
    user = User.objects.get(username=username)
    mu_pos = MapUserPosition.objects.get(current_project=project, user=user)
    data = {
        'latlng': mu_pos.latlng.split(','),
        'zoom': mu_pos.zoom
    }
    return Response(data)


@api_view(['POST', ])
def save_user_position(request):
    project = request.project    
    zoom = request.POST.get('zoom')
    latlng = request.POST.get('latlng')
    username = request.COOKIES.get('username')
    user = User.objects.get(username=username)
    if MapUserPosition.objects.filter(current_project=project, user=user).exists():
        mu_pos =  MapUserPosition.objects.get(current_project=project, user=user)
        mu_pos.latlng = latlng
        mu_pos.zoom = zoom
        mu_pos.save()
    else:
        MapUserPosition.objects.create(
            current_project=project, 
            user=user,
            latlng = latlng,
            zoom = zoom)
    return Response([])


@api_view(['POST', ])
def save_map_position(request):
    project = request.project    
    zoom = request.POST.get('zoom')
    latlng = request.POST.get('latlng')
    username = request.COOKIES.get('username')
    if UserSettings.objects.filter(current_project=project, user=username).exists():
        us = UserSettings.objects.get(current_project=project, user=username)        
        us.map_zoom = zoom
        us.map_center=latlng
        us.save()        
    else:        
        UserSettings.objects.create(
            current_project=project,
            user=username,
            gsm_file = '',
            wcdma_file = '',
            lte_file = '',
            rnd_gsm_file = '',
            rnd_wcdma_file = '',
            rnd_lte_file = '',
            gsm_color = '#ffa500',
            wcdma_color = '#0000FF',
            lte_color = '#008000',
            element_color = '#FF0000',
            gsm_radius = '1000',
            wcdma_radius = '1200',
            lte_radius = '1500',
            map_center = '0,0',
            map_zoom = '10',
            co_pci_color = '#FF0000',
            sc_color = '#FF0000',
            adj_minus_color = '#00FF00',
            adj_plus_color = '#0000FF',
            co_bcch_color = '#FF0000',
            deleted_neighbour_color = '#0000FF',
            new_neighbour_color = '#ffa500',
            neighbour_color = '#FF0000',
            selected_cell_color = '#00FF00'
        )
    return Response([])


@api_view(['GET', ])
def get_param_values(request, network, param):
    project = request.project
    values = Rnd(project.id, network).get_param_values(param)
    return Response(values)


@api_view(['GET', ])
def get_rnd_neighbors(request, network, sector):
    project = request.project
    filename = request.wcdma.filename
    values = Rnd(project.id, network).get_rnd_neighbors(sector, project.id)
    return Response(values)


@api_view(['GET', ])
def get_new3g(request, network, sector):
    project = request.project
    filename = request.wcdma.filename
    values = Rnd(project.id, network).get_new3g(sector, filename)
    return Response(values)


@api_view(['POST', ])
def new3g3g(request):
    project = request.project
    if request.POST:
        Rnd(project.id, 'wcdma').save_new_3g(
            request.wcdma.filename,
            request.POST.get('rncSource'),
            request.POST.get('utrancellSource'),
            request.POST.get('carrierSource'),
            request.POST.get('rncTarget'),
            request.POST.get('utrancellTarget'),
            request.POST.get('carrierTarget'),
            request.POST.get('status'))

    return Response([])


@api_view(['POST', ])
def del3g3g(request):
    project = request.project
    if request.POST:
        Rnd(project.id, 'wcdma').del_3g(
            request.wcdma.filename,
            request.POST.get('utrancellSource'),
            request.POST.get('utrancellTarget'))
    return Response([])


@api_view(['POST', ])
def flush3g3g(request):
    project = request.project
    if request.method == 'POST':
        Rnd(project.id, 'wcdma').flush_3g(request.wcdma.filename)
    return Response([])


def get3g3gscript(request):
    project = request.project
    filename = Rnd(project.id, 'wcdma').create_script(request.wcdma.filename)
    return HttpResponseRedirect(filename)


@api_view(['GET', ])
def get_rnd_pd(request, network, sector, date_from, date_to):
    project = request.project
    pd = Distance().get_chart(project.id, date_from, date_to, sector)
    return Response(pd)


@api_view(['GET', ])
def get_same_neighbor(request):
    project = request.project
    return Response(Rnd(project.id, 'wcdma').same_neighbor(request.wcdma.filename))


@api_view(['POST', ])
@gzip_page
def drive_test(request):
    project_id = request.project.id
    map_bounds = request.POST.get('bounds').split(',')
    zoom = int(request.POST.get('zoom'))
    ms = request.POST.get('ms')
    param = request.POST.get('parameter')
    filenames = request.POST.get('filenames').split(',')
    legend = []
    for p in DriveTestLegend.objects.filter(project=request.project, param=request.POST.get('legend')):
        legend.append({
            'param': p.param,
            'min_value': p.min_value,
            'max_value': p.max_value,
            'color': p.color
        })

    dt = DriveTest()
    data = dt.get_points(project_id, filenames, ms, param, legend, map_bounds, zoom)
    return Response(data)


@api_view(['GET', ])
def drive_test_init(request):
    project = request.project
    project_id = request.project.id
    dt = DriveTest()
    legends = []
    data = dt.init_drive_test(project_id)
    for l in DriveTestLegend.objects.filter(project=project):
        legends.append({'param': l.param, 'max_val': l.max_value, 'min_val': l.min_value, 'color': l.color})

    data['legends'] = list(set(l.get('param') for l in legends))
    data['full_legends'] = legends
    return Response(data)


def set_drive_test_legend(request):
    project = request.project
    filename = handle_uploaded_file(request.FILES.getlist('uploaded_file'))[0]
    wb = load_workbook(filename=filename, use_iterators=True)
    result = []
    sheet_name = wb.get_sheet_names()[0]
    ws = wb.get_sheet_by_name(sheet_name)
    param_name = ''
    DriveTestLegend.objects.filter(project=project).delete()
    for row in ws.iter_rows():
        if row[0].value == '**':
            param_name = row[1].value
        else:
            if row[3].value:
                if row[1].value:
                    min_val = row[1].value
                else:
                    min_val = ''
                if row[2].value:
                    max_val = row[2].value
                else:
                    max_val = ''
                color = row[3].value

            result.append([param_name, min_val, max_val, color])

            DriveTestLegend.objects.create(
                project=project,
                param=param_name,
                min_value=min_val,
                max_value=max_val,
                color=color)
    return HttpResponse(json.dumps(result), content_type='application/json')


@api_view(['GET', ])
def get_drive_test_legend(request):
    project = request.project
    data = []
    for dtl in DriveTestLegend.objects.filter(project=project):
        data.append([dtl.param, dtl.min_value, dtl.max_value, dtl.color])
    return Response(data)


@api_view(['GET', ])
def get_drive_test_param(request, ms):
    dt = DriveTest()
    project = request.project
    params = dt.get_params(project.id, ms)
    return Response(params)


@api_view(['GET', ])
def drive_test_point(request, id):
    dt = DriveTest()
    project = request.project
    point = dt.get_point(project.id, id)
    return Response(point)
    
@gzip_page
@api_view(['GET', ])
def universal_table(request, relation):
    project = request.project
    print project.id
    ut = UniversalTable(relation.lower())
    columns, data = ut.get_table(project.id)
    if request.GET.get('excel'):
        return HttpResponseRedirect(Excel(request.project.project_name, relation, columns, data).filename)

    return Response({'columns': columns, 'data': data})

@api_view(['GET', ])
@gzip_page
def measurements_wncs(request):
    project = request.project
    distance = request.GET.get('distance')
    drop = request.GET.get('drop')
    cells = request.GET.get('cells')
    excel = request.GET.get('excel')
    params = dict(project=project)
    if (distance != 'undefined') and distance:
        params['distance__lte'] = distance
    if (drop != 'undefined') and drop:
        params['drop__gte'] = drop
    if (cells != 'undefined') and cells:
        params['cell_name__icontains'] = cells

    wncs = WNCS.objects.filter(**params).order_by('-drop')
    data = []
    for row in wncs:
        data.append({
            'id': row.id,
            'cell_name': row.cell_name,
            'nb_cell_name': row.nb_cell_name,
            'sc': row.sc,
            'events': row.events,
            'drop': row.drop,
            'distance': row.distance
        })
    if excel == 'true':
        excel_data = []
        for row in data:
            excel_data.append([
                row.get('cell_name'),
                row.get('nb_cell_name'),
                row.get('sc'),
                row.get('events'),
                row.get('drop'),
                row.get('distance')
            ])
        s = Excel(
            project.project_name,
            'wncs',
            ['CellName', 'nbCellName', 'SC', 'Events', 'Drop call', 'Distance[km]'],
            excel_data).filename
        print s
        return HttpResponseRedirect(s)
    else:
        return Response({'data': data})


@api_view(['GET', ])
@gzip_page
def measurements_wncs_top(request):
    project = request.project
    distance = request.GET.get('distance')
    drop = request.GET.get('drop')
    cells = request.GET.get('cells')
    params = dict(project=project)
    if (distance != 'undefined') and distance:
        params['distance__lte'] = distance
    if (drop != 'undefined') and drop:
        params['drop__gte'] = drop
    if (cells != 'undefined') and cells:
        params['cell_name__icontains'] = cells

    wncs = WNCS.objects.filter(**params).order_by('-drop')[:20]
    data = []
    for row in wncs:
        data.append({
            'cell_name': row.cell_name,
            'drop': row.drop,
        })
    return Response({'data': data})
    
@api_view(['POST', ])
def rnd_from_network_data(request):
    project = request.project
    rnd_name = request.POST.get('rnd_name')
    network = request.POST.get('network')
    UniversalTable('').create_rnd_from_network(project, network, rnd_name)
    return Response({'data': []})
    

