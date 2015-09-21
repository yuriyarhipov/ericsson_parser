import json

from django.db import connection
from openpyxl import load_workbook


class Rnd:

    def __init__(self, project_id, network):
        self.project_id = project_id
        self.network = network

        cursor = connection.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS rnd (
                id SERIAL,
                project_id INT,
                network TEXT,
                data JSON)
            ''')
        cursor.close()
        connection.commit()

    def get_data(self):
        cursor = connection.cursor()
        cursor.execute('''
            SELECT
                data
            FROM
                rnd
            WHERE
                (project_id=%s) AND (network=%s)''', (
            self.project_id,
            self.network))

        result = cursor.fetchone()[0]
        return result

    def write_file(self, filename):
        cursor = connection.cursor()
        wb = load_workbook(filename=filename, use_iterators=True)
        ws = wb.active
        columns = []
        data = []
        column_row = True
        for excel_row in ws.iter_rows():
            if column_row:
                columns = [cell.value for cell in excel_row]
            else:
                row = dict()
                for column in columns:
                    row[column] = excel_row[columns.index(column)].value
                data.append(row)
            column_row = False
        result = dict(columns=columns, data=data)

        cursor.execute('''
            DELETE FROM
                rnd
            WHERE
                (project_id=%s) AND (network=%s)''', (
            self.project_id,
            self.network))

        cursor.execute('''
            INSERT INTO
                rnd (
                    project_id,
                    network,
                    data)
            VALUES (%s, %s, %s)''', (
            self.project_id,
            self.network,
            json.dumps(result, encoding='latin1')
        ))
        connection.commit()

        return result

