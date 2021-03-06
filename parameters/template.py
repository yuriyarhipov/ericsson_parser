from collections import OrderedDict
from openpyxl import load_workbook

from files.models import Files
from query.models import QueryTemplate, SiteQuery
from django.conf import settings
from collections import OrderedDict
import psycopg2


class Template(object):

    mo = []
    params = []
    template_name = ''
    tables = dict()

    def __init__(self, out_of_range=True):
        self.conn = psycopg2.connect(
            'host = %s dbname = %s user = %s password = %s' % (
                settings.DATABASES['default']['HOST'],
                settings.DATABASES['default']['NAME'],
                settings.DATABASES['default']['USER'],
                settings.DATABASES['default']['PASSWORD']))
        self.cursor = self.conn.cursor()
        self.out_of_range = out_of_range

    def save_template(self, project, network, template_name, mo, params, min_values, max_values):
        self.template_name = template_name
        self.network = network
        self.project = project
        QueryTemplate.objects.filter(template_name=template_name).delete()
        for i in range(0, int(len(mo))):
            table_name = mo[i]
            param_name = params[i] if len(params) > i else ''
            min_value = min_values[i] if len(min_values) > i else ''
            max_value = max_values[i] if len(max_values) > i else ''

            if table_name and param_name:
                qt = QueryTemplate()
                qt.project = project
                qt.network = network
                qt.mo = table_name
                qt.param_name = param_name
                qt.max_value = max_value
                qt.min_value = min_value
                qt.template_name = self.template_name
                qt.save()

    def get_columns(self, table_name):
        cursor = self.conn.cursor()
        cursor.execute("SELECT column_name FROM information_schema.columns WHERE lower(table_name)='%s';" % table_name.lower())
        return [r[0].lower() for r in cursor]

    def get_join_lte(self, table_name, filename, cells):
        columns = self.get_columns(table_name)
        sql_cells = ','.join(cells)
        sql_key = ''
        topology_key = ''
        if 'eutrancell' in columns:
            sql_key = 'EUtrancell'
            topology_key = 'EUtrancell'
        elif 'element1' in columns:
            sql_key = 'Element1'
            topology_key = 'CID'
        elif 'element2' in columns:
            sql_key = 'Element2'
            topology_key = 'SITE'
        print 'A'
        print cells
        if cells:
            sql = 'TOPOLOGY_LTE INNER JOIN %s ON ((%s.%s = TOPOLOGY_LTE.%s) AND (%s.filename=TOPOLOGY_LTE.filename) AND (%s.filename IN (%s)) AND (TOPOLOGY_LTE.EUtrancell in (%s)))' % (
                table_name,
                table_name,
                sql_key,
                topology_key,
                table_name,
                table_name,
                filename,
                sql_cells)
        else:
            sql = 'TOPOLOGY_LTE INNER JOIN %s ON ((%s.%s = TOPOLOGY_LTE.%s) AND (%s.filename=TOPOLOGY_LTE.filename) AND (%s.filename IN (%s)))' % (
                table_name,
                table_name,
                sql_key,
                topology_key,
                table_name,
                table_name,
                filename)

        return sql

    def get_join_wcdma(self, table_name, filename, cells):
        columns = self.get_columns(table_name)
        sql_cells = ','.join(cells)
        join_sql = []
        if 'utrancell' in columns:
            join_sql.append('(%s.UtranCell = TOPOLOGY.UtranCell)' % table_name)
        if 'element1' in columns:
            join_sql.append('(%s.Element1 = TOPOLOGY.RNC)' % table_name)
        elif 'element2' in columns:
            join_sql.append('(%s.Element2 = TOPOLOGY.SITE)' % table_name)
        if 'sectorcarrier' in columns:
            join_sql.append('(%s.SectorCarrier = TOPOLOGY.SectorCarrier)' % table_name)
        if 'carrier' in columns:
            join_sql.append('(%s.Carrier = TOPOLOGY.Carrier)' % table_name)
        join_sql = ' AND '.join(join_sql)

        if cells:
            sql = 'TOPOLOGY INNER JOIN %s ON (%s AND (TOPOLOGY.filename=%s.filename) AND (%s.filename IN (%s)) AND (Topology.UtranCell IN (%s)) )' % (
                table_name,
                join_sql,
                table_name,
                table_name,
                filename,
                sql_cells)
        else:
            sql = 'TOPOLOGY INNER JOIN %s ON (%s AND (TOPOLOGY.filename=%s.filename) AND (%s.filename IN (%s)))' % (
                table_name,
                join_sql,
                table_name,
                table_name,
                filename)
        return sql

    def get_tables_cna(self, sql_tables, filename, cells):
        tables = []
        result_columns = []
        sql_cells = ','.join(cells)
        cell_index = 1
        for table_name, columns in sql_tables.iteritems():
            tables.append(table_name)
            sql_columns = []
            for column in columns:
                if column.lower() == 'cell':
                    sql_columns.append('"%s"."%s" AS CELL_%s' % (table_name, column, str(cell_index)))
                    cell_index += 1
                else:
                    sql_columns.append('"%s"."%s"' % (table_name, column))
        result_columns = result_columns + sql_columns

        sql_tables = ''
        root_table = tables[0]
        if len(tables) == 1:
            sql_tables = '"%s"' % tables[0]
        else:
            temp_tables = []
            for table in tables[1:]:
                temp_tables.append('INNER JOIN "%s" ON "%s".CELL="%s".CELL' % (table, root_table, table))
                sql_tables = '%s %s' % (root_table, ' '.join(temp_tables))

        sql_columns = ','.join(result_columns)
        sql = '''SELECT DISTINCT "%s".CELL, %s FROM %s WHERE ("%s".CELL IN (%s)) AND ("%s".filename='%s')''' % (
            root_table,
            sql_columns,
            sql_tables,
            root_table,
            sql_cells,
            root_table,
            filename)
        return sql

    def get_tables_lte(self, sql_tables, filename, cells):
        tables = []
        result_columns = []
        for table_name, columns in sql_tables.iteritems():
            sql_columns = ','.join(['%s.%s' % (table_name, col) for col in columns])
            sql_join = self.get_join_lte(table_name,filename, cells)
            table_sql = 'INNER JOIN (SELECT DISTINCT TOPOLOGY_LTE.EUtrancell, TOPOLOGY_LTE.filename, %s FROM %s) AS T_%s ON ((TOPOLOGY_LTE.EUtrancell=T_%s.EUtrancell) AND (TOPOLOGY_LTE.filename=T_%s.filename))' % (sql_columns, sql_join, table_name, table_name, table_name)
            tables.append(table_sql)
            result_columns.extend(['T_%s.%s' % (table_name, col) for col in columns])

        sql_columns = ', '.join(result_columns)
        sql_join = ' '.join(tables)
        sql = 'SELECT DISTINCT TOPOLOGY_LTE.CID, TOPOLOGY_LTE.SITE, TOPOLOGY_LTE.EUtrancell, TOPOLOGY_LTE.filename, %s FROM TOPOLOGY_LTE %s' % (sql_columns, sql_join)
        return sql

    def get_tables_wcdma(self, sql_tables, filename, cells):
        tables = []
        result_columns = []
        for table_name, columns in sql_tables.iteritems():
            if table_name and columns:
                sql_columns = ','.join(['%s.%s' % (table_name, col) for col in columns])
                sql_join = self.get_join_wcdma(table_name, filename, cells)
                table_sql = 'INNER JOIN (SELECT DISTINCT Topology.UtranCell, Topology.filename, %s FROM %s) AS T_%s ON ((Topology.Utrancell=T_%s.Utrancell) AND (Topology.filename=T_%s.filename))' % (sql_columns, sql_join, table_name, table_name, table_name)
                tables.append(table_sql)
                result_columns.extend(['T_%s.%s' % (table_name, col) for col in columns])

        sql_columns = ', '.join(result_columns)
        sql_join = ' '.join(tables)
        sql = 'SELECT TOPOLOGY.RNC, TOPOLOGY.UTRANCELL, TOPOLOGY.SITE, Topology.SectorCarrier, Topology.SectorAntena, Topology.Carrier, TOPOLOGY.filename,  %s  FROM Topology %s' % (sql_columns, sql_join)
        return sql

    def get_tables(self, sql_tables, network, filename, cells):
        if network == 'WCDMA':
            return self.get_tables_wcdma(sql_tables, filename, cells)
        elif network == 'LTE':
            return self.get_tables_lte(sql_tables, filename, cells)
        elif network == 'GSM':
            return self.get_tables_cna(sql_tables, filename, cells)

    def get_table_from_column(self, column_name, file_tables):
        cursor = self.conn.cursor()
        cursor.execute("SELECT DISTINCT table_name FROM INFORMATION_SCHEMA.COLUMNS WHERE (lower(column_name)='%s')" % (column_name.lower(), ))
        tables = []
        for row in cursor:
            tables.append(row[0].lower())

        if 'utrancell' in tables:
            tables = ['utrancell', ]
        for table in tables:
            if ('template_' not in table) and (table.lower() in file_tables):
                cursor.execute("SELECT column_name FROM information_schema.columns WHERE lower(table_name)='%s';" % table.lower())
                columns = [r[0] for r in cursor]
                if (column_name.lower() in columns) and (('utrancell' in columns) or ('element1' in columns) or ('element2' in columns) or ('cell')):
                    return table

    def get_select(self, project, template_name, filename, cells):
        sql_tables = OrderedDict()
        network = 'WCDMA'
        clear_filename = filename.replace("'", '')
        file_columns = set(t.lower() for t in Files.objects.filter(filename=clear_filename).first().tables.split(','))
        for template in QueryTemplate.objects.filter(template_name=template_name, project=project).order_by('id'):
            table_name = self.get_table_from_column(template.param_name, file_columns)
            if table_name:
                column = template.param_name
                network = template.network
                if table_name not in sql_tables:
                    sql_tables[table_name] = []
                if column not in sql_tables[table_name]:
                    sql_tables[table_name].append(column)
        sql = self.get_tables(sql_tables, network, filename, cells)
        return sql

    def create_indexes(self, template_name):
        qt = QueryTemplate.objects.filter(template_name=template_name).first()
        network = qt.network
        cursor = self.conn.cursor()
        try:
            if network == 'WCDMA':
                cursor.execute('CREATE INDEX ON "template_%s" (filename);' % template_name)
                cursor.execute('CREATE INDEX ON "template_%s" (utrancell);' % template_name)
                self.conn.commit()
        except:
            return

    def get_sql_compare_id(self, filename):
        f = Files.objects.filter(filename=filename).first()
        if f:
            if f.network == 'WCDMA':
                return 'utrancell'
            elif f.network == 'LTE':
                return 'eutrancell'
        return 'CELL'

    def get_site_query_value(self, filename, column, value, param, table):
        cursor = self.conn.cursor()
        cursor.execute("SELECT %s FROM %s WHERE (filename='%s') AND (lower(%s)='%s')" % (param, table, filename, column, value.lower()))
        if cursor.rowcount > 0:
            return cursor.fetchall()[0][0]

    def check_columns(self, column, table_name):
        cursor = self.conn.cursor()
        cursor.execute("SELECT table_name FROM INFORMATION_SCHEMA.COLUMNS WHERE (lower(column_name)=%s) AND (lower(table_name)=%s)", (column.lower(), table_name.lower()))
        return cursor.rowcount > 0

    def get_element2_value(self, filename, utrancell):
        cursor = self.conn.cursor()
        cursor.execute("SELECT Site FROM Topology WHERE (filename='%s') AND (lower(Utrancell)=lower('%s'))" % (filename, utrancell))
        for row in cursor:
            return row[0]

    def check_value(self, value, p_min, p_max):
        status = 'default'
        if p_min and value:
            value = value.strip()
            p_min = p_min.strip()
            p_max = p_max.strip()
            if p_min.lower() in ['true', 'false']:
                if value:
                    status = 'success'
                else:
                    status = 'danger'

            elif (float(value) >= float(p_min)) and (float(value) <= float(p_max)):
                status = 'success'
            else:
                status = 'danger'
        return [value, status]

    def get_site_query(self, site, filename):
        sourcefile = Files.objects.filter(filename=filename).first()
        params = sourcefile.get_site_query(site)
        return params

    def site_query(self, project, network, filename):
        data = OrderedDict()
        wb = load_workbook(filename=filename, use_iterators=True)
        ws = wb.active
        SiteQuery.objects.filter(network=network).delete()
        for row in ws.iter_rows():
            if row[0].value == '**':
                site_name = row[1].value
            else:
                param_name = row[1].value
                param_min = row[2].value if row[2].value else ''
                param_max = row[3].value if row[3].value else ''
                if param_name:
                    if site_name not in data:
                        data[site_name] = []
                    data[site_name].append([param_name, param_min, param_max])
                    SiteQuery.objects.create(
                        project=project,
                        network=network,
                        site=site_name,
                        param_name=param_name,
                        param_min=param_min,
                        param_max=param_max)
        return data

    def get_sites(self, filename):
        data = []
        cursor = self.conn.cursor()

        source_file = Files.objects.filter(filename=filename).first()
        if source_file:
            if source_file.network == 'WCDMA':
                cursor.execute("SELECT DISTINCT Utrancell FROM Topology WHERE (filename='%s')" % (filename, ))
                for row in cursor:
                    data.append(dict(name=row[0]))
            elif source_file.network == 'LTE':
                cursor.execute("SELECT DISTINCT EUtrancell FROM Topology_LTE WHERE (filename='%s')" % (filename, ))
                for row in cursor:
                    data.append(dict(name=row[0]))
            elif source_file.network == 'GSM':
                cursor.execute("SELECT DISTINCT CELL FROM COMMON_CELL_DATA WHERE (filename='%s')" % (filename, ))
                for row in cursor:
                    data.append(dict(name=row[0]))
        return data

    def get_table(self, project, table_name, cells, column, min_value, max_value):        
        cursor = self.conn.cursor()
        sql_columns = [column, ]
        if min_value and max_value and self.out_of_range:
            sql_columns.append(min_value + ' "Min:"')
            sql_columns.append(max_value + '"Max:"')
     
        cursor.execute('SELECT column_name FROM INFORMATION_SCHEMA.COLUMNS WHERE LOWER(table_name)=%s', (table_name.lower(), ))
        columns = [row[0] for row in cursor]
        where_sql = []        
        sql_cells = ','.join(["'%s'" % cell for cell in cells])
        if 'utrancell' in columns:
            sql_columns.append('Utrancell')
            where_sql.append('Utrancell not in (%s)' % sql_cells)            
        if 'element2' in columns:
            sql_columns.append('Element2')
            where_sql.append('Element2 not in (%s)' % sql_cells)
        if 'element1' in columns:
            sql_columns.append('Element1')
            where_sql.append('Element1 not in (%s)' % sql_cells)
        where_sql = ' AND '.join(where_sql)
                                    
        sql_columns = reversed(sql_columns)
        if min_value and max_value and self.out_of_range:            
            try:                
                cursor.execute(''' SELECT ''' + ','.join(sql_columns) + ''' FROM ''' + table_name + ''' WHERE (project_id::integer=%s) AND NOT ((''' + column + '''::float>=%s) AND (''' + column + '''::float<=%s)) AND NOT (''' + where_sql +''')''', ( project.id, float(min_value), float(max_value)))
            except:                
                cursor.close()
                self.conn.rollback()
                cursor = self.conn.cursor()
        else:            
            try:                                                
                cursor.execute(''' SELECT ''' + ','.join(sql_columns) + ''' FROM ''' + table_name + ''' WHERE (project_id::integer=%s)  AND NOT (''' + where_sql +''')''', ( project.id, ))
            except:                
                cursor.close()
                self.conn.rollback()
                cursor = self.conn.cursor()
        columns = []
        data = []
        if cursor.rowcount > 0:
            columns = ['%s' % desc[0] for desc in cursor.description]
                        
            for row in cursor:
                data_row = dict()
                for col in columns:
                    data_row[col] = row[columns.index(col)]
                data.append(data_row)                     
        return columns, data

    def get_parameter(self, project, cells, param_name, min_value, max_value):
        cursor = self.conn.cursor()        
        cursor.execute('''
            SELECT
                table_name
            FROM
                INFORMATION_SCHEMA.COLUMNS
            WHERE
                (lower(column_name)=%s)
            ORDER BY table_name;
            ''', (param_name.lower(), ))
        tabs = OrderedDict()
        if cursor.rowcount == 0:            
            tabs[param_name] = {'columns': [], 'data': []}
        tables = [row[0].upper() for row in cursor]
        if 'UTRANCELL' in tables:
            columns, data = self.get_table(project, 'Utrancell', cells, param_name, min_value, max_value)
            tabs['%s (%s)(%s)' % (param_name, 'Utrancell', len(data))] = {'columns': columns, 'data': data}
        else:
            for table_name in tables:
                if table_name not in ['topology', ]:                
                    columns, data = self.get_table(project, table_name, cells, param_name, min_value, max_value)
                    tabs['%s (%s)(%s)' % (param_name, table_name, len(data))] = {'columns': columns, 'data': data}
        
        
        return tabs

    def get_data(self, project, template, cells):
        data = OrderedDict()
        for t in QueryTemplate.objects.filter(project=project, template_name=template).order_by('id'):            
            param_data = self.get_parameter(project, cells, t.param_name, t.min_value, t.max_value)
            for key, value in param_data.iteritems():                
                data[key] = value 
        return data
