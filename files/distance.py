import json
from openpyxl import load_workbook
from django.db import connection
from os.path import basename
from datetime import datetime
from decimal import Decimal
from files.models import Files


class Distance(object):

    def __init__(self):
        cursor = connection.cursor()

        cursor.execute('''CREATE TABLE IF NOT EXISTS Access_Distance (
            project_id INT,
            filename TEXT,
            RNC  TEXT,
            RBS TEXT,
            date_id date,
            sector INT,
            DCVECTOR_INDEX BIGINT,
            pmPropagationDelay BIGINT,
            Carrier BIGINT,
            Utrancell TEXT,
            Distance REAL)
            ''')
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS LogicalSector (
                id SERIAL,
                project_id INT,
                logical_sector TEXT,
                Technology TEXT,
                Band INT,
                SECTOR TEXT,
                Label TEXT)''')
        try:
            cursor.execute('''
                CREATE INDEX
                    utrancell_date_idx
                ON
                    Access_Distance (Utrancell, date_id);''')
            cursor.execute('''
                CREATE INDEX
                    utrancell_idx
                ON
                    Access_Distance (Utrancell);''')
        except:
            pass
        connection.commit()

    def delete_file(self, filename, project_id):
        cursor = connection.cursor()
        cursor.execute('DELETE FROM Access_Distance WHERE (filename=%s) AND (project_id=%s)',(
            filename,
            project_id
        ))

    def get_sectors(self):
        cursor = connection.cursor()
        cursor.execute('SELECT DISTINCT Utrancell FROM Access_Distance')
        sectors = [row[0] for row in cursor]
        return sectors

    def get_rbs(self, project_id):
        cursor = connection.cursor()
        cursor.execute('SELECT DISTINCT RBS FROM Access_Distance WHERE project_id=%s ORDER BY RBS', (project_id, ))
        rbs = [row[0] for row in cursor]
        return rbs

    def get_dates(self, rbs):
        cursor = connection.cursor()
        cursor.execute("SELECT DISTINCT date_id FROM Access_Distance WHERE (RBS='%s') ORDER BY date_id" % rbs)
        dates = [row[0].strftime('%d.%m.%Y') for row in cursor]
        return dates

    def get_low_coverage(self, day_from, day_to, distance, project_id):
        cursor = connection.cursor()
        cursor.execute('''
            SELECT
                DIST_SUM.RNC,
                DIST_SUM.Utrancell,
                DIST_SUM.dist_sum,
                DATE_SUM.date_sum
            FROM (
            SELECT
                RNC,
                SUM(pmpropagationdelay) AS dist_sum,
                Utrancell
            FROM
                Access_Distance
            WHERE
                (date_id >=%s) AND (date_id <=%s) AND (DCVECTOR_INDEX <= %s)
            GROUP BY
                RNC,
                Utrancell
            ) AS DIST_SUM,
            (
            SELECT
                sum(pmpropagationdelay) date_sum,
                Utrancell,
                RNC
            FROM
                Access_Distance
            WHERE
                (date_id >=%s) AND (date_id <=%s)
            GROUP BY
                Utrancell,
                RNC
            ) AS date_sum

            WHERE (DATE_SUM.Utrancell=DIST_SUM.Utrancell) AND (DATE_SUM.RNC=DIST_SUM.RNC)
            ORDER BY DIST_SUM.RNC, DIST_SUM.Utrancell
            ''', (
            datetime.strptime(day_from, '%d.%m.%Y'),
            datetime.strptime(day_to, '%d.%m.%Y'),
            distance,
            datetime.strptime(day_from, '%d.%m.%Y'),
            datetime.strptime(day_to, '%d.%m.%Y')))

        data = []
        for row in cursor:
            data.append({
                'rnc': row[0],
                'utrancell': row[1],
                'dist_sum': int(row[2]),
                'date_sum': int(row[3])})
        return data

    def get_over_coverage(self, day_from, day_to, distance, project_id):
        cursor = connection.cursor()
        cursor.execute('''
            SELECT
                DIST_SUM.RNC,
                DIST_SUM.Utrancell,
                DIST_SUM.dist_sum,
                DATE_SUM.date_sum
            FROM (
            SELECT
                RNC,
                SUM(pmpropagationdelay) AS dist_sum,
                Utrancell
            FROM
                Access_Distance
            WHERE
                (date_id >=%s) AND (date_id <=%s) AND (DCVECTOR_INDEX > %s)
            GROUP BY
                RNC,
                Utrancell
            ) AS DIST_SUM,
            (
            SELECT
                sum(pmpropagationdelay) date_sum,
                Utrancell,
                RNC
            FROM
                Access_Distance
            WHERE
                (date_id >=%s) AND (date_id <=%s)
            GROUP BY
                Utrancell,
                RNC
            ) AS date_sum

            WHERE (DATE_SUM.Utrancell=DIST_SUM.Utrancell) AND (DATE_SUM.RNC=DIST_SUM.RNC)
            ORDER BY DIST_SUM.RNC, DIST_SUM.Utrancell
            ''', (
            datetime.strptime(day_from, '%d.%m.%Y'),
            datetime.strptime(day_to, '%d.%m.%Y'),
            distance,
            datetime.strptime(day_from, '%d.%m.%Y'),
            datetime.strptime(day_to, '%d.%m.%Y')))

        data = []
        for row in cursor:
            data.append({
                'rnc': row[0],
                'utrancell': row[1],
                'dist_sum': int(row[2]),
                'date_sum': int(row[3])})
        return data

    def get_charts(self, day_from, day_to, rbs, project_id):
        cursor = connection.cursor()
        cursor.execute('''
            SELECT DISTINCT
                Utrancell
            FROM
                Access_Distance
            WHERE
                (RBS=%s) AND (project_id=%s)''', (
            rbs,
            project_id))

        data = dict()
        for row in cursor:
            utrancell = row[0]
            chart, title, distances = self.get_chart(day_from, day_to, utrancell)
            data[utrancell] = dict(
                distances=distances,
                chart=chart,
                title=title)
        return data

    def get_chart(self, date_from, date_to, utrancell):
        data = []
        cursor = connection.cursor()

        cursor.execute('''
                SELECT
                    DIST_SUM.distance,
                    DIST_SUM.dist_sum,
                    DIST_SUM.DCVECTOR_INDEX,
                    DATE_SUM.date_sum,
                    Sector,
                    Carrier
                FROM (
                SELECT
                    distance,
                    SUM(pmpropagationdelay) AS dist_sum,
                    DCVECTOR_INDEX,
                    Sector,
                    Carrier
                FROM
                    Access_Distance
                WHERE
                    (Utrancell=%s) AND (date_id >=%s) AND (date_id <=%s)
                GROUP BY
                    distance,
                    DCVECTOR_INDEX,
                    Sector,
                    Carrier
                ) AS DIST_SUM,
                (
                SELECT
                    sum(pmpropagationdelay) date_sum
                FROM
                    Access_Distance
                WHERE
                    (Utrancell=%s) AND (date_id >=%s) AND (date_id <=%s)
                ) AS date_sum ORDER BY DIST_SUM.distance
                ''', (
                    utrancell,
                    datetime.strptime(date_from, '%d.%m.%Y'),
                    datetime.strptime(date_to, '%d.%m.%Y'),
                    utrancell,
                    datetime.strptime(date_from, '%d.%m.%Y'),
                    datetime.strptime(date_to, '%d.%m.%Y')))



        distances = dict()
        for row in cursor:
            value = Decimal(row[1]) / Decimal(row[3]) * 100
            distances[int(row[2])] = float(row[0])
            data.append([
                int(row[2]),
                round(value, 2), ]
            )
            carrier = row[4]
            sector = row[5]

        title = '''
            <b>Utrancell:</b>%s
            <b>Carrier:</b>%s
            <b>Sector:</b>%s''' % (
            utrancell,
            carrier,
            sector)
        return data, title, distances

    def write_file(self, project, description, vendor, filename, current_task):
        wb = load_workbook(filename=filename, use_iterators=True)
        ws = wb.active
        cursor = connection.cursor()
        cursor.execute('DELETE FROM Access_Distance WHERE filename=%s',
                       (basename(filename), ))
        i = 0
        rows = []

        for excel_row in ws.iter_rows():
            row = []
            if i > 0:
                for cell in excel_row:
                    if cell.value:
                        row.append(cell.value)
                    else:
                        row.append(0)
                rows.append(row)

                if len(rows) == 10000:
                    sql_rows = []
                    for row in rows:
                        sql_rows.append(cursor.mogrify(
                            '(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)',
                            (
                                project.id,
                                basename(filename),
                                row[0],
                                row[1],
                                row[2],
                                row[3],
                                row[4],
                                row[5],
                                row[6],
                                row[7],
                                row[8])))
                    cursor.execute('''
                        INSERT INTO
                            Access_Distance
                        VALUES
                            %s''' % ','.join(sql_rows))
                    rows = []
            i += 1

            if i % 10000 == 0:
                current_task.update_state(state="PROGRESS", meta={
                    "current": int(i / 10000)})
        sql_rows = []
        for row in rows:
            sql_rows.append(cursor.mogrify(
                '(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)',
                (
                    project.id,
                    basename(filename),
                    row[0],
                    row[1],
                    row[2],
                    row[3],
                    row[4],
                    row[5],
                    row[6],
                    row[7],
                    row[8])))
        cursor.execute('''
            INSERT INTO
                Access_Distance
            VALUES
                %s''' % ','.join(sql_rows))

        connection.commit()
        Files.objects.filter(filename=basename(filename), project=project).delete()
        Files.objects.create(
            filename=basename(filename),
            file_type='HISTOGRAM FILE COUNTER - Access Distance',
            project=project,
            tables='',
            description=description,
            vendor=vendor,
            network='WCDMA')

    def get_logical_distr(self, project_id, day_from, day_to, rbs, sectors):
        if not sectors:
            return []
        cursor = connection.cursor()
        data = []
        utrancells = ["'%s%s'" % (rbs, s) for s in sectors]
        cursor.execute('''
            SELECT
                utrancell,
                SUM(pmpropagationdelay)
            FROM
                access_distance
            WHERE
                (utrancell IN (%s)) AND
                (project_id = %s) AND
                (date_id >='%s') AND (date_id <='%s')
            GROUP BY
                utrancell
            ORDER BY utrancell;''' % (
            ','.join(utrancells),
            project_id,
            datetime.strptime(day_from, '%d.%m.%Y'), datetime.strptime(day_to, '%d.%m.%Y')))


        for row in cursor:
            data.append(dict(name=row[0], y=int(row[1])))

        return data

    def get_distr(self, day_from, day_to, rbs, project_id):
        data = []
        cursor = connection.cursor()
        cursor.execute('''
            SELECT
                MIN(date_id),
                MAX(date_id)
            FROM
                Access_Distance
            WHERE
                (RBS=%s)''', (
            rbs, ))
        days = cursor.fetchall()[0]

        logical_sectors = dict()
        cursor.execute('SELECT logical_sector, technology, sector  FROM LogicalSector WHERE project_id=%s', (project_id, ))

        for row in cursor:
            logical_sector = row[0]
            if logical_sector not in logical_sectors:
                logical_sectors[logical_sector] = []
            if row[1] == 'WCDMA':
                logical_sectors[logical_sector].append(row[2])

        for ls_name, ls_sectors in logical_sectors.iteritems():
            data.append({
                'logical_sector': ls_name,
                'data': self.get_logical_distr(project_id, day_from, day_to, rbs, ls_sectors)
            })

        return data

    def logical_sectors(self, project_id):
        cursor = connection.cursor()
        result = []
        cursor.execute('SELECT logical_sector, technology, band, sector, Label FROM LogicalSector WHERE project_id=%s', (project_id, ))
        for row in cursor:
            result.append({
                'logical_sector': row[0],
                'technology': row[1],
                'band': row[2],
                'sector': row[3],
                'label': row[4]})
        return result

    def add_logical_sector(self, project_id, logical_sector, technology, band, sector):
        cursor = connection.cursor()
        cursor.execute('''
                INSERT INTO LogicalSector (project_id, logical_sector, technology, band, sector, label)
                VALUES (%s, %s, %s, %s, %s, %s)''', (
            project_id,
            logical_sector,
            technology,
            band,
            sector,
            '%s %s-%s' % (sector, technology, band)
        ))
        connection.commit()

    def delete_logical_sectors(self, project_id, logical_sector, sector):
        cursor = connection.cursor()
        cursor.execute('DELETE FROM LogicalSector WHERE (project_id=%s) AND (logical_sector=%s) AND (sector=%s)', (
            project_id,
            logical_sector,
            sector))
        connection.commit()
