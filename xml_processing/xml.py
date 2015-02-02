import re
import psycopg2
import tempfile
import xlsxwriter
from openpyxl import load_workbook

from lxml import etree
from os.path import basename
from celery import current_task
from os.path import join
from zipfile import ZipFile

from django.conf import settings
from files.models import Files
from tables.table import Topology
from files.tasks import parse_xml


def get_mo(mo):
    result = dict()
    pattern = '(\w*)=(\w*-*\w*)'
    k = re.compile(pattern)
    for k, v in re.findall(k, mo):
        result[k] = v
    return result


class LoadExcel:
    tables = dict()
    data = dict()
    filename = ''

    def __init__(self, filename):
        self.wb = load_workbook(filename=filename, use_iterators=True)
        self.main()

    def get_data(self, table_name):
        result = []
        data = []
        ws = self.wb.get_sheet_by_name(name=table_name)
        for row in ws.iter_rows():
            r = []
            for cell in row:
                r.append(cell.internal_value)
            data.append(r)
        count = len(data[0])
        for row in data[1:]:
            r = dict()
            for i in range(count - 1):
                r[data[0][i]] = row[i]
                if data[0][i] == 'FileName':
                    self.filename = row[i]
            result.append(r)

        return result

    def get_columns(self, table_name):
        ws = self.wb.get_sheet_by_name(name=table_name)
        return set(cell.internal_value for cell in ws.iter_rows().next())


    def main(self):
        table_names = self.wb.get_sheet_names()
        self.tables = {table_name: self.get_columns(table_name) for table_name in table_names}
        for table in table_names:
            self.data[table] = self.get_data(table)


class Excel:
    static_path = settings.STATICFILES_DIRS[0]
    workbook = None
    archive_filename = ''
    filename = ''

    def __init__(self, filename):
        self.conn = psycopg2.connect(
            'host = %s dbname = %s user = %s password = %s' % (
                settings.DATABASES['default']['HOST'],
                settings.DATABASES['default']['NAME'],
                settings.DATABASES['default']['USER'],
                settings.DATABASES['default']['PASSWORD']))
        self.cursor = self.conn.cursor()
        self.filename = filename
        self.main()

    def write_field_to_xls(self, worksheet, row, field):
        i = 0
        for column in field:
            worksheet.write(row, i, column)
            i += 1

    def write_table_to_xls(self, table_name):
        try:
            worksheet = self.workbook.add_worksheet(table_name)
            i = 0
            self.cursor.execute("SELECT DISTINCT * FROM %s" % table_name)
            columns = [desc[0].lower() for desc in self.cursor.description]
            sql = "SELECT %s FROM %s WHERE filename='%s';" % (','.join(columns), table_name, self.filename)
            self.cursor.execute(sql)
            for column in columns:
                worksheet.write(0, i, column)
                i += 1

            i = 1
            for field in self.cursor:
                if i > 65530:
                    break
                self.write_field_to_xls(worksheet, i, field)
                i += 1
        except:
            pass


    def main(self):
        task = current_task
        path = tempfile.mkdtemp()
        excel_filename = join(path, 'Export_MultiTable.xlsx')
        self.workbook = xlsxwriter.Workbook(excel_filename, {'constant_memory': True})
        tables, suffix = [[], '']  # displayed_tables(self.filename)
        count = len(tables)
        i = 0
        for table_name in tables:
            i += 1
            self.write_table_to_xls(table_name)
            task.update_state(state='PROGRESS',
                              meta={'current': 50 + int((((float(i) / float(count)) * 99) / 2)), 'total': 99})

        self.workbook.close()
        self.archive_filename = join(self.static_path, self.filename[:-4] + '.zip')

        zip = ZipFile(self.archive_filename, 'w')
        zip.write(excel_filename, arcname='Export_MultiTable.xlsx')
        zip.close()
        del self.workbook
        del zip
        self.cursor.close()
        self.conn.close()
        task.update_state(state='PROGRESS',
                          meta={'current': 99, 'total': 99})


class Tables:
    def __init__(self, data, tables, filename):
        self.data = data
        self.tables = tables
        self.filename = filename
        self.file_type = 'WCDMA' if 'UtranCell' in data.keys() else 'LTE'
        self.conn = psycopg2.connect(
            'host = %s dbname = %s user = %s password = %s' % (
                settings.DATABASES['default']['HOST'],
                settings.DATABASES['default']['NAME'],
                settings.DATABASES['default']['USER'],
                settings.DATABASES['default']['PASSWORD']))
        self.cursor = self.conn.cursor()

    def check_table(self, table_name, columns):
        exists_columns = set([c.lower() for c in self.tables[table_name]])
        missed_columns = exists_columns - columns
        for column in missed_columns:
            self.cursor.execute('ALTER TABLE %s ADD COLUMN %s text;' % (table_name, column))
        self.conn.commit()
        self.add_indexes(table_name, missed_columns)

    def load_data(self, task, current, interval_per_file):
        count = len(self.tables)
        interval = float(interval_per_file) / float(count)
        for table_name, source_columns in self.tables.iteritems():
            columns = []
            filtred_columns = []
            for col in source_columns:
                if col.lower() not in filtred_columns:
                    columns.append(col)
                    filtred_columns.append(col.lower())
            current = float(current) + float(interval)
            task.update_state(state="PROGRESS", meta={"current": int(current), "total": 99})
            values = []
            for field in self.data[table_name]:
                row = dict()
                for column in columns:
                    row[column] = field.get(column, '')
                values.append(row)
            values_name = ['%(' + column + ')s' for column in columns]
            insert = 'INSERT INTO %s (%s) VALUES (%s)' % (table_name, ','.join(columns), ','.join(values_name))
            self.cursor.executemany(insert, values)

    def add_indexes(self, table_name, columns):
        for column in columns:
            if column.lower() in ['utrancell', 'element1', 'element2']:
                self.cursor.execute('CREATE INDEX ON %s (%s)' % (table_name, column))
        self.conn.commit()

    def create_table(self, table_name):
        columns = []
        for column in self.tables[table_name]:
            if column.lower() not in columns:
                columns.append(column.lower())

        sql_columns = ['%s TEXT' % field for field in columns]
        self.cursor.execute('CREATE TABLE IF NOT EXISTS %s (%s);' % (table_name, ', '.join(sql_columns)))
        self.conn.commit()
        self.add_indexes(table_name, self.tables[table_name])

    def table(self, table_name):
        self.cursor.execute(
            'SELECT column_name FROM information_schema.columns WHERE table_catalog = %s AND table_name=%s;',
            (settings.DATABASES['default']['NAME'], table_name.lower()))
        column_names = set(row[0].lower() for row in self.cursor)
        if column_names:
            self.check_table(table_name, column_names)
            self.cursor.execute('DELETE FROM ' + table_name + ' WHERE filename=%s;', (self.filename, ))
        else:
            self.create_table(table_name)

    def fourgneighbors(self):
        if 'EUtrancellFDD'.lower() not in [t.lower() for t in self.tables]:
            return

        self.cursor.execute('DROP TABLE IF EXISTS FourGNeighbors;')
        self.cursor.execute('''
            SELECT DISTINCT
                filename,
	            EUtrancellFDD Source,
	            target,
	            cellIndividualOffsetEUtran,
	            coverageIndicator,
	            createdBy,
	            includeInSystemInformation,
	            isHoAllowed,
	            isRemoveAllowed,
	            lastModification,
	            loadBalancing,
	            qOffsetCellEUtran
		    INTO FourGNeighbors
            FROM EUtranCellRelation
            ;''')

    def threegneighborss(self):
        self.cursor.execute('DROP TABLE IF EXISTS ThreeGNeighbors;')
        self.cursor.execute('''
          SELECT DISTINCT
	        Utrancell as Source,
	        Neighbor as Target,
	        SelectionPriority as Priority,
	        frequencyRelationType,
	        CASE
		      WHEN frequencyRelationType = '1' THEN 'INTER_FREQ'
		      ELSE 'INTRA_FREQ'
	        END
	        as RelationType,
	        nodeRelationType,
	        CASE
		      WHEN frequencyRelationType = '1' THEN 'INTER_RNC'
		      ELSE 'INTRA_RNC'
	        END
	        as node_RelationType,
	        loadSharingCandidate,
	        filename
	        INTO ThreeGNeighbors
            FROM
              UtranRelation
            WHERE
              (Neighbor IS NOT NUll) AND
              (Utrancell IS NOT NULL);
        ''')

    def topology(self):
        self.cursor.execute('DROP TABLE IF EXISTS TOPOLOGY CASCADE;')
        self.cursor.execute('''
            SELECT DISTINCT
	          UtranCell.Element1 RNC,
	          RBSLocalCell.Element2 SITE,
	          UtranCell.UtranCell,
	          Utrancell.CID,
	          substring(RBSLocalCell.SectorCarrier from 2 for 1) Sector,
	          RBSLocalCell.SectorCarrier SectorCarrier,
	          substring(RBSLocalCell.SectorCarrier from 2 for 1) SectorAntena,
	          substring(RBSLocalCell.SectorCarrier from 4 for 1) Carrier,
	          IubLink.Iublink,
	          UtranCell.filename
            INTO TOPOLOGY
            FROM RBSLocalCell
              INNER JOIN UtranCell ON (RBSLocalCell.LocalCellid = UtranCell.CID and RBSLocalCell.filename = UtranCell.filename)
              LEFT JOIN IubLink ON  (RBSLocalCell.Element2 = IubLink.Element2 AND RBSLocalCell.filename = IubLink.filename)
            ;''')

    def create_topology_tree_view(self):
        cursor = self.conn.cursor()
        self.cursor.execute('CREATE TABLE IF NOT EXISTS TOPOLOGY_TREEVIEW (FILENAME TEXT, TREEVIEW JSON);')
        self.conn.commit()
        cursor.execute('SELECT DISTINCT filename FROM TOPOLOGY')
        for row in cursor:
            Topology(row[0]).create_tree_view()

    def topology_lte(self):
        tables = ','.join(self.tables.keys())
        if 'eutrancellfdd' not in tables.lower():
            return

        self.cursor.execute('DROP TABLE IF EXISTS TOPOLOGY_LTE;')
        self.cursor.execute('''
            SELECT
                EUtrancellFDD.filename,
                EUtrancellFDD.Element2 Site,
                EUtrancellFDD.EUtrancellFDD EUtrancell,
                CellId CID,
                physicalLayerCellIdGroup PHYSICAL_CID,
                SectorEquipmentFunction.SectorEquipmentFunction,
                SectorEquipmentFunction.AntennaUnitGroup,
                TPTM1.mmeName mmeName_1,
                TPTM2.mmeName mmeName_2
            INTO TOPOLOGY_LTE
            FROM EUtrancellFDD
	            INNER JOIN SectorEquipmentFunction ON SectorEquipmentFunction.EUtrancellFDD=EUtrancellFDD.EUtrancellFDD
	            INNER JOIN TermPointToMme TPTM1 ON TPTM1.Element2=EUtrancellFDD.Element2
	            INNER JOIN TermPointToMme TPTM2 ON TPTM2.Element2=EUtrancellFDD.Element2
            WHERE
	            (TPTM1.TermPointToMme='1') AND
	            (TPTM2.TermPointToMme='2') AND
	            (EUtrancellFDD.filename=SectorEquipmentFunction.filename) AND
	            (EUtrancellFDD.filename=TPTM1.filename) AND
	            (EUtrancellFDD.filename=TPTM2.filename)
            ;''')

    def rnd_wcdma(self):
        self.cursor.execute('DROP VIEW IF EXISTS BrightcommsRNDDate;')
        self.cursor.execute('DROP TABLE IF EXISTS RND_WCDMA;')
        self.cursor.execute('''
            SELECT DISTINCT
                topology.filename,
                Sector.MO,
                Sector.element2 as site,
                Sector.sector,
                topology.cid,
                topology.utrancell,
                topology.carrier,
                topology.rnc,
                latitude,
                longitude,
                beamdirection,
                height,
                band,
                lathemisphere,
                geodatum,
                MEContext.ipaddress,
                MEContext.nemimversion,
                MEContext.mirrormibversion,
                MEContext.lostsynchronisation,
                ManagedElement.userdefinedstate,
                ManagedElement.vendorname,
                ManagedElement.swversion,
                ManagedElement.productname,
                ManagedElement.siteref,
                ManagedElement.logicalname,
                UtranCell.uarfcnul,
                UtranCell.uarfcndl,
                UtranCell.primaryscramblingcode,
                UtranCell.primarycpichpower,
                UtranCell.lac,
                UtranCell.rac,
                UtranCell.sac,
                UtranCell.cellreserved,
                UtranCell.administrativeState
            INTO
                RND_WCDMA
            FROM
                Sector
                INNER JOIN Topology ON ((Sector.element2=Topology.site) AND (Sector.sector=Topology.sector))
                INNER JOIN MEContext ON (Sector.element2=MEContext.element2)
                INNER JOIN ManagedElement ON (Sector.element2=ManagedElement.element2)
                INNER JOIN UtranCell ON ((topology.cid=UtranCell.cid))
            WHERE (Sector.filename=Topology.filename) AND
                (Sector.filename=MEContext.filename) AND
                (Sector.filename=ManagedElement.filename) AND
                (Sector.filename=UtranCell.filename)
            ;''')


    def rnd_lte(self):
        tables = ','.join(self.tables.keys())
        if 'eutrancellfdd' not in tables.lower():
            return
        self.cursor.execute('DROP TABLE IF EXISTS RND_LTE;')
        self.cursor.execute('''
            SELECT
                EUtrancellFDD.filename,
                EUtrancellFDD.Element2 Site,
                EUtrancellFDD.EUtrancellFDD,
                SectorEquipmentFunction.SectorEquipmentFunction,
                SectorEquipmentFunction.AntennaUnitGroup,
                EUtrancellFDD.latitude,
                EUtrancellFDD.longitude,
                SectorEquipmentFunction.fqband,
                EUtrancellFDD.ulChannelBandwidth,
                EUtrancellFDD.dlChannelBandwidth,
                EUtrancellFDD.administrativestate,
                EUtrancellFDD.earfcndl,
                EUtrancellFDD.earfcnul,
                MeContext.ipaddress,
                MeContext.neMIMversion,
                MeContext.lostSynchronisation,
                ManagedElement.userDefinedState,
                ManagedElement.managedElementType,
                ManagedElement.vendorname,
                ManagedElement.swversion,
                ManagedElement.productname,
                ManagedElement.siteref,
                EUtrancellFDD.physicalLayerCellIdGroup,
                EUtrancellFDD.noOfTxAntennas,
                EUtrancellFDD.noOfRxAntennas,
                EUtrancellFDD.cellbarred,
                EUtrancellFDD.qrxlevmin,
                EUtrancellFDD.qrxlevminOffset,
                EUtrancellFDD.qQualMin,
                EUtrancellFDD.qQualMinOffset,
                EUtrancellFDD.maximumTransmissionPower,
                SectorEquipmentFunction.confOutputPower,
                SectorEquipmentFunction.sectorPower,
                AntennaSubunit.totalTilt,
                Antennaunit.mechanicalAntennaTilt
            INTO RND_LTE
            FROM EUtrancellFDD
	            INNER JOIN SectorEquipmentFunction ON SectorEquipmentFunction.EUtrancellFDD=EUtrancellFDD.EUtrancellFDD
	            INNER JOIN MeContext ON MeContext.Element2=EUtrancellFDD.Element2
	            INNER JOIN ManagedElement ON ManagedElement.Element2=EUtrancellFDD.Element2
	            INNER JOIN AntennaSubunit ON AntennaSubunit.Element2=EUtrancellFDD.Element2
	            INNER JOIN Antennaunit ON Antennaunit.Element2=EUtrancellFDD.Element2
            WHERE
	            (EUtrancellFDD.filename=SectorEquipmentFunction.filename) AND
	            (EUtrancellFDD.filename=MeContext.filename) AND
	            (EUtrancellFDD.filename=ManagedElement.filename) AND
	            (EUtrancellFDD.filename=AntennaSubunit.filename) AND
	            (EUtrancellFDD.filename=Antennaunit.filename)
            ;''')

    def create_tables(self, task, current, interval_per_file):
        for table_name in self.tables:
            self.table(table_name)
        self.load_data(task, current, interval_per_file)
        self.topology()
        self.topology_lte()
        self.rnd_wcdma()
        self.rnd_lte()
        self.fourgneighbors()
        self.threegneighborss()
        self.conn.commit()
        self.create_topology_tree_view()
        self.cursor.close()
        self.conn.close()
        del self.tables
        del self.data


class Processing:
    data = dict()
    tables = dict()
    ignored_name = []
    version = ''
    vendorName = ''
    path = re.compile('\{.*\}')
    ver = re.compile('^\D+')
    stop_list = ['ManagementNode', ]

    def __init__(self, filename):
        self.xml_file = filename
        self.filename = basename(filename)

    def get_table_name(self, node):
        table_name = None
        parent = node.getparent()
        if 'VsDataContainer' in parent.tag:
            data = node.find(".//{genericNrm.xsd}vsDataType")
            if data is not None:
                table_name = data.text[6:]
        else:
            table_name = self.path.sub('', parent.tag)
        return table_name

    def get_additional_fields(self, node, table_name):
        result = dict()
        parent = node.getparent()
        if ('VsDataContainer' in parent.tag) and (table_name in parent.getparent().tag):
            for n in parent.getparent():
                if 'attributes' in n.tag:
                    for attr in n:
                        field_name = self.path.sub('', attr.tag)
                        text = getattr(attr, 'text', '')
                        if text is None:
                            text = ''
                        text = text.replace('vsData', '')
                        result[field_name] = text
        return result

    def get_fields(self, node, table_name):
        result = dict(
            FileName=basename(self.filename),
            MO=', '.join(self.get_mo(node)),
            vendorName=self.vendorName,
            version=self.version)
        result.update(self.get_additional_fields(node, table_name))

        mo = get_mo(result['MO'])

        if 'UtranCell' in mo:
            result['UtranCell'] = mo.get('UtranCell', '')

        if 'EUtranCellFDD' in mo:
            result['EUtranCellFDD'] = mo.get('EUtranCellFDD', '')

        if 'SectorEquipmentFunction' in mo:
            result['SectorEquipmentFunction'] = mo.get('SectorEquipmentFunction', '')

        if 'AntennaUnitGroup' in mo:
            result['AntennaUnitGroup'] = mo.get('AntennaUnitGroup', '')

        if table_name == 'GsmRelation':
            result['GsmRelation'] = mo.get('GsmRelation', '')

        if 'IubLink' in mo:
            result['IubLink'] = mo.get('IubLink')

        if 'Iub' in mo:
            result['Iub'] = mo.get('Iub')

        if 'IurLink' in mo:
            result['IurLink'] = mo.get('IurLink')

        if 'UeRabType' in mo:
            result['UeRabType'] = mo.get('UeRabType')

        if 'UeRc' in mo:
            result['UeRc'] = mo.get('UeRc')

        if 'Carrier' in mo:
            result['Carrier'] = mo.get('Carrier')

        if 'TermPointToMme' in mo:
            result['TermPointToMme'] = mo.get('TermPointToMme')

        if 'Sector' in mo:
            if 'Element' not in result:
                result['Element'] = mo.get('MeContext', '')
            result['Sector'] = mo.get('Sector')

        if 'RbsLocalCell' in mo:
            if 'Element' not in result:
                result['Element'] = mo.get('MeContext', '')
            result['SectorCarrier'] = mo.get('RbsLocalCell')

        site = mo.get('MeContext')
        sub = mo.get('SubNetwork')

        if site and sub:
            if site == sub:
                result['Element1'] = site
            else:
                result['Element2'] = site

        for n in node.iter():
            if 'vsDataFormatVersion' in n.tag:
                result['version'] = self.ver.sub('', n.text)
                self.version = result['version']
            if n.text is None:
                n.text = ''

            parent = n.getparent()
            if ('{genericNrm.xsd}attributes' not in parent.tag) and ('vsdata' not in parent.tag.lower()):
                field_name = '%s_%s' % (self.path.sub('', parent.tag), self.path.sub('', n.tag))
            else:
                field_name = self.path.sub('', n.tag)
            text = n.text.strip().replace('vsData', '')
            if ('vsData' not in field_name) and (field_name not in ['attributes']):
                if field_name in result:
                    result[field_name] = '%s %s' % (result[field_name], text)
                else:
                    result[field_name] = text

        if table_name == 'UtranRelation':
            if 'adjacentCell' in result:
                ac = get_mo(result.get('adjacentCell'))
                result['Neighbor'] = ac.get('UtranCell')
        elif table_name == 'IubLink':
            if 'iubLinkNodeBFunction' in result:
                ib = get_mo(result.get('iubLinkNodeBFunction'))
                result['Element2'] = ib.get('MeContext')

        elif table_name == 'SectorEquipmentFunction':
            rb = get_mo(result.get('reservedBy'))
            rf_branch_ref = get_mo(result.get('rfBranchRef'))
            result['EUtranCellFDD'] = rb.get('EUtranCellFDD')
            result['AntennaUnitGroup'] = rf_branch_ref.get('AntennaUnitGroup')

        elif table_name == 'EUtranCellRelation':
            ac = get_mo(result.get('adjacentCell'))
            result['Target'] = ac.get('EUtranCellFDD')

        elif table_name == 'CoverageRelation':
            tc = get_mo(result.get('utranCellRef'))
            result['Target_coverage'] = tc.get('UtranCell', '')

        return result

    def get_mo(self, node):
        result = []
        parent = node.getparent()
        if parent is not None:
            result = self.get_mo(parent)
        value = node.get('id')
        name = None
        if value is not None:
            tag = self.path.sub('', node.tag)
            if tag == 'VsDataContainer':
                datatype = node.find(".//{genericNrm.xsd}vsDataType")
                if datatype is not None:
                    name = datatype.text[6:]
            else:
                name = tag
            if name and value:
                result.append('%s=%s' % (name, value))
        return result

    def get_table(self, attribute):
        table_name = self.get_table_name(attribute)
        if table_name in self.stop_list:
            return
        if table_name and (table_name not in self.data):
            self.data[table_name] = []
            self.tables[table_name] = set()

        fields = self.get_fields(attribute, table_name)

        columns = self.tables[table_name]
        self.data[table_name].append(fields)
        if columns:
            self.tables[table_name] = columns | set(fields.keys())
        else:
            self.tables[table_name] = set(fields.keys())

    def parse_data(self, task, current, interval_per_file):
        tree = etree.parse(self.xml_file)
        root = tree.getroot()

        header = root.find('{configData.xsd}fileHeader')
        if header is not None:
            self.vendorName = header.get('vendorName')

        for attribute in root.iterfind('.//{genericNrm.xsd}attributes'):
            self.get_table(attribute)


class Xml(object):
    def save_xml(self, filename, project, description, vendor, file_type, network, current_task, current,
                 interval_per_file):
        xml = Processing(filename)
        xml.parse_data(current_task, current, interval_per_file / 2)
        tables = Tables(xml.data, xml.tables, xml.filename)
        tables.create_tables(current_task, current, interval_per_file / 2)
        Files.objects.filter(filename=xml.filename, project=project).delete()
        Files.objects.create(
            filename=xml.filename,
            file_type=file_type,
            project=project,
            tables=','.join(xml.tables.keys()),
            description=description,
            vendor=vendor,
            network=network)


