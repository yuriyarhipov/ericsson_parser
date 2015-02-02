import json

from openpyxl import load_workbook

from django.http import HttpResponse

from query.models import GroupCells
from files.cna import CNA
from files.views import handle_uploaded_file


def get_cells(request, technology):
    data = []
    if (technology == 'WCDMA') and request.wcdma:
        data = request.wcdma.get_cells()
    elif (technology == 'GSM') and request.gsm:
        data = CNA().get_cells(request.gsm.filename)

    return HttpResponse(json.dumps(data), content_type='application/json')


def save_group_of_cells(request):
    project = request.project
    title = request.POST.get('title')
    network = request.POST.get('network')
    cells = ','.join(request.POST.getlist('cells[]'))
    GroupCells.objects.create(project=project, group_name=title, network=network, cells=cells)
    return HttpResponse(json.dumps({'success': 'ok'}), content_type='application/json')


def get_groups(request):
    project = request.project
    data = []
    for group in GroupCells.objects.filter(project=project):
        data.append({'name': group.group_name, 'network': group.network})
    return HttpResponse(json.dumps(data), content_type='application/json')

def upload_cells_template(request):
    data =[]
    filename = handle_uploaded_file(request.FILES.getlist('excel'))[0]
    wb = load_workbook(filename=filename, use_iterators=True)
    ws = wb.active
    for row in ws.iter_rows():
        data.append({'cell': row[0].value, 'type': 'Cells'})
    return HttpResponse(json.dumps(data), content_type='application/json')
