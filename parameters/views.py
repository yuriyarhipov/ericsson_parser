import json
from collections import OrderedDict

from openpyxl import load_workbook

from django.http import HttpResponse
from django.shortcuts import HttpResponseRedirect
from django.db import connection
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.conf import settings

import xlsxwriter
from zipfile import ZipFile
from os.path import join
import tempfile

from query.models import QueryTemplate, SiteQuery
from files.wcdma import WCDMA
from files.cna import CNA
from files.lte import LTE
from files.views import handle_uploaded_file
from files.models import Files
from template import Template
from tables.table import get_excel
from files import tasks


def version_release(request):
    cursor = connection.cursor()
    columns = ['version', 'name']
    data = []

    cursor.execute("select DISTINCT version, vendorName from UtranCell")
    for row in cursor:
        data.append([row[0], '%s WCDMA' % row[1], ])

    return HttpResponse(json.dumps({'columns': columns, 'data': data}), content_type='application/json')


def get_param(request, network):
    data = []
    if network == 'WCDMA':
        data = request.wcdma.get_param()
    elif network == 'GSM':
        data = request.gsm.get_param()
    elif network == 'LTE':
        data = request.lte.get_param()

    return HttpResponse(json.dumps(data), content_type='application/json')


def add_template(request):
    data = []
    param_values = request.POST.getlist('param')
    min_values = request.POST.getlist('min_value')
    max_values = request.POST.getlist('min_value')
    template_name = request.POST.get('template_name')
    project = request.project
    network = request.POST.get('network')
    QueryTemplate.objects.filter(project=request.project, template_name=template_name).delete()
    for i in range(0, len(param_values)):
        QueryTemplate.objects.create(
            project=project,
            network=network,
            template_name=template_name,
            param_name=param_values[i],
            min_value=min_values[i],
            max_value=max_values[i]
            )

    #if network == 'GSM':
    #    CNA().create_template(template_name)
    #elif network == 'WCDMA':
    #    tasks.create_parameters_table.delay(request.wcdma, network, template_name)
    #elif network == 'LTE':
    #    Template().create_template_table(request.lte, template_name)


    return HttpResponse(json.dumps(data), content_type='application/json')


def predefined_templates(request):
    data = []
    for qt in QueryTemplate.objects.filter(project=request.project).distinct('template_name').order_by('template_name'):
        data.append({'template_name': qt.template_name, 'network': qt.network})
    return HttpResponse(json.dumps(data), content_type='application/json')


def delete_template(request, template_name):
    QueryTemplate.objects.filter(project=request.project, template_name=template_name).delete()
    return HttpResponse(json.dumps({'sucess': 'ok', }), content_type='application/json')


def get_templates(request):
    templates = [qt.template_name for qt in QueryTemplate.objects.filter(project=request.project).distinct('template_name')]
    return HttpResponse(json.dumps(templates), content_type='application/json')


def get_template_cells(request, network, filename):
    data = []
    if network == 'GSM':
        data = CNA().get_cells(filename)
    elif network == 'WCDMA':
        wcdma = WCDMA()
        data = wcdma.get_cells(filename)
    elif network == 'LTE':
        data = LTE().get_cells(filename)

    return HttpResponse(json.dumps(data), content_type='application/json')


def run_template(request):
    project = request.project
    template = request.GET.get('template')
    tabs = Template().get_data(project, template)

    if request.GET.get('excel'):
        static_path = settings.STATICFILES_DIRS[0]
        archive_filename = join(static_path, template +'.zip')
        excel_filename = join(tempfile.mkdtemp(), template + '.xlsx')
        workbook = xlsxwriter.Workbook(excel_filename, {'constant_memory': True})        
        for worksheet_name in tabs:
            columns = tabs.get(worksheet_name).get('columns')            
            worksheet = workbook.add_worksheet(columns[-1][:30])
            i = 0
            for column in tabs.get(worksheet_name).get('columns'):               
                worksheet.write(0, i, column)
                i += 1
            row_id = 1            
            for row in tabs.get(worksheet_name).get('data'):                
                i = 0
                for col in columns:                    
                    worksheet.write(row_id, i, row.get(col))
                    i += 1
                row_id += 1
                    
        workbook.close()
        zip = ZipFile(archive_filename, 'w')
        zip.write(excel_filename, arcname=template + '.xlsx')
        zip.close()
        excel_data = []        
        return HttpResponseRedirect('/static/%s' % get_excel(template, columns, excel_data))
    return HttpResponse(json.dumps(tabs), content_type='application/json')


def upload_template(request):
    data =[]
    filename = handle_uploaded_file(request.FILES.getlist('excel'))[0]
    wb = load_workbook(filename=filename, use_iterators=True)
    ws = wb.active
    for row in ws.iter_rows():
        if row[0].value != 'PARAMETER':
            excel_param = row[0].value
            if excel_param:
                data.append(dict(param=excel_param, min_value=row[1].value, max_value=row[2].value))

    return HttpResponse(json.dumps(data), content_type='application/json')


def edit_template(request, template):
    data = {}
    data['template'] = template
    data['param_table'] = []
    for qt in QueryTemplate.objects.filter(template_name=template).all():
        data['network'] = qt.network
        data['param_table'].append({'mo': qt.mo, 'param': qt.param_name, 'min_value': qt.min_value, 'max_value': qt.max_value})
    return HttpResponse(json.dumps(data), content_type='application/json')


def save_automatic_site_query(request):
    network = request.POST.get('network')
    filename = handle_uploaded_file(request.FILES.getlist('uploaded_file'))[0]
    data = Template().site_query(project=request.project, network=network, filename=filename)
    return HttpResponse(json.dumps({'data': data}), content_type='application/json')


def automatic_site_query(request, network):
    data = OrderedDict()
    for query in SiteQuery.objects.filter(project=request.project, network=network):
        if query.site not in data:
            data[query.site] = []
        data[query.site].append([query.param_name, query.param_min, query.param_max])
    return HttpResponse(json.dumps({'data': data}), content_type='application/json')


def get_site_query(request, network, site):
    data = dict()
    data['tabs'] = []
    if network == 'WCDMA':
        source_file = request.wcdma.filename
    elif network == 'LTE':
        source_file = request.lte.filename
    elif network == 'GSM':
        source_file = request.gsm.filename

    params = Template().get_site_query(site, source_file)
    for tab_name, tab_params in params.iteritems():
        data['tabs'].append({'title': tab_name, 'content': tab_params})
    return HttpResponse(json.dumps(data), content_type='application/json')


def get_sites(request, network):
    if network == 'WCDMA':
        source_file = request.wcdma.filename
    elif network == 'LTE':
        source_file = request.lte.filename
    elif network == 'GSM':
        source_file = request.gsm.filename
    data = Template().get_sites(source_file)
    return HttpResponse(json.dumps(data), content_type='application/json')


def get_network_files(request, network):
    data = dict()
    data['files'] = [f.filename for f in Files.objects.filter(project=request.project, network=network)]
    return HttpResponse(json.dumps(data), content_type='application/json')
